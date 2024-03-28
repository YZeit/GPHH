import operator
import numpy as np
import streamlit as st
import pandas as pd
import simpy
import matplotlib as mpl
import matplotlib.pyplot as plt
import openpyxl as op
import networkx as nx
import random
from deap import algorithms
from deap import base
from deap import creator
from deap import tools
from deap import gp
from statistics import mean
from PIL import Image
from simulation import simulation
import multiprocessing as mp
import matplotlib.ticker as mtick
from math import pi


def dominates(X1, X2):
    if(np.any(X1 < X2) and np.all(X1 <= X2)):
        return True
    else:
        return False

def is_pareto_efficient_dumb(costs):
    """
    Find the pareto-efficient points
    :param costs: An (n_points, n_costs) array
    :return: A (n_points, ) boolean array, indicating whether each point is Pareto efficient
    """
    is_efficient = np.ones(costs.shape[0], dtype = bool)
    for i, c in enumerate(costs):
        is_efficient[i] = np.all(np.any(costs[:i]>c, axis=1)) and np.all(np.any(costs[i+1:]>c, axis=1))
    return is_efficient

def keep_efficient(pts):
    'returns Pareto efficient row subset of pts'
    # sort points by decreasing sum of coordinates
    pts = pts[pts.sum(1).argsort()[::-1]]
    # initialize a boolean mask for undominated points
    # to avoid creating copies each iteration
    undominated = np.ones(pts.shape[0], dtype=bool)
    for i in range(pts.shape[0]):
        # process each point in turn
        n = pts.shape[0]
        if i >= n:
            break
        # find all points not dominated by i
        # since points are sorted by coordinate sum
        # i cannot dominate any points in 1,...,i-1
        undominated[i+1:n] = (pts[i+1:] >= pts[i]).any(1)
        # keep points undominated so far
        pts = pts[undominated[:n]]
    return pts

def create_statistics_visualization(nb_generation, avgFitnessValues, minFitnessValues, maxFitnessValues):
    # plot statistics:
    fig, ax = plt.subplots()
    # create stacked errorbars:
    ax.errorbar(nb_generation, avgFitnessValues, [avgFitnessValues - minFitnessValues, maxFitnessValues - avgFitnessValues],
                 fmt='.k', ecolor='gray', lw=1)
    ax.set_xlabel('Generation')
    ax.set_ylabel('Average Fitness')
    st.header('Evolution process')
    return fig

def create_visualization_zoom(reference_point, solution, factor):

    reference_point_intesification = [0, 0, 0]
    reference_point_intesification[0] = reference_point[0] + factor * (
            solution[0] - reference_point[0])
    reference_point_intesification[1] = reference_point[1] + factor * (
            solution[1] - reference_point[1])
    reference_point_intesification[2] = reference_point[2] + factor * (
            solution[2] - reference_point[2])
    fig2 = plt.figure()
    ax2 = fig2.add_subplot(projection='3d')
    ax2.scatter(solution[0], solution[1], solution[2], label='selected solution', color = 'g')
    ax2.scatter(reference_point[0], reference_point[1],
                reference_point[2], label='original reference point', color = 'b')
    ax2.scatter(reference_point_intesification[0], reference_point_intesification[1],
                reference_point_intesification[2], label='new reference point', color = 'y')
    ax2.plot([solution[0], reference_point[0]],
             [solution[1], reference_point[1]],
             [solution[2], reference_point[2]], color='b', linewidth=1)

    ax2.plot([solution[0], reference_point[0]], [solution[1], solution[1]],
             [solution[2], solution[2]], color='r', dashes=[6, 2], linewidth=1)
    ax2.plot([reference_point[0], reference_point[0]], [solution[1], reference_point[1]],
             [solution[2], solution[2]], color='r', dashes=[6, 2], linewidth=1)
    ax2.plot([reference_point[0], reference_point[0]], [solution[1], solution[1]],
             [solution[2], reference_point[2]], color='r', dashes=[6, 2], linewidth=1)
    ax2.plot([reference_point[0], reference_point[0]], [reference_point[1], reference_point[1]],
             [solution[2], reference_point[2]], color='r', dashes=[6, 2], linewidth=1)
    ax2.plot([solution[0], solution[0]], [reference_point[1], reference_point[1]],
             [solution[2], reference_point[2]], color='r', dashes=[6, 2], linewidth=1)
    ax2.plot([solution[0], solution[0]], [solution[1], reference_point[1]],
             [reference_point[2], reference_point[2]], color='r', dashes=[6, 2], linewidth=1)
    ax2.plot([reference_point[0], reference_point[0]], [solution[1], reference_point[1]],
             [reference_point[2], reference_point[2]], color='r', dashes=[6, 2], linewidth=1)
    ax2.plot([solution[0], reference_point[0]], [reference_point[1], reference_point[1]],
             [reference_point[2], reference_point[2]], color='r', dashes=[6, 2], linewidth=1)
    ax2.plot([solution[0], reference_point[0]], [reference_point[1], reference_point[1]],
             [solution[2], solution[2]], color='r', dashes=[6, 2], linewidth=1)
    ax2.plot([solution[0], reference_point[0]], [solution[1], solution[1]],
             [reference_point[2], reference_point[2]], color='r', dashes=[6, 2], linewidth=1)
    ax2.plot([solution[0], solution[0]], [solution[1], solution[1]],
             [solution[2], reference_point[2]], color='r', dashes=[6, 2], linewidth=1)
    ax2.plot([solution[0], solution[0]], [solution[1], reference_point[1]],
             [solution[2], solution[2]], color='r', dashes=[6, 2], linewidth=1)

    ax2.set_xlabel('Makespan', fontsize=8)
    ax2.set_ylabel('Total tardiness', fontsize=8)
    ax2.set_zlabel('Total waiting time', fontsize=8)
    ax2.tick_params(axis='both', which='major', labelsize=6)
    ax2.tick_params(axis='both', which='minor', labelsize=8)
    ax2.view_init(20, 15)
    ax2.legend(loc='best', bbox_to_anchor=(1, 1), fontsize='x-small')
    return fig2, reference_point_intesification

def create_statistics_visualization_boxplot(nb_generation, avgFitnessValues, minFitnessValues, maxFitnessValues, stdFitnessValues):
    # plot statistics:
    fig, ax = plt.subplots()
    # create stacked errorbars:
    ax.errorbar(nb_generation, avgFitnessValues, [avgFitnessValues - minFitnessValues, maxFitnessValues - avgFitnessValues],
                 fmt='.k', ecolor='gray', lw=1)
    ax.set_xlabel('Generation')
    ax.set_ylabel('Average Fitness')
    st.header('Evolution process')
    return fig

def create_relative_performance(makespan, tardiness, waiting, x_labels):
    labels = [i for i in range(1,len(makespan)+1)]
    x = np.arange(len(makespan))
    width = 0.25
    fig, ax = plt.subplots()
    rects1 = ax.bar(x - width , makespan*100, width, label='Makespan')
    rects2 = ax.bar(x, tardiness*100, width, label='Total Tardiness')
    rects3 = ax.bar(x + width , waiting*100, width, label='Total waiting time')
    # Add some text for labels, title and custom x-axis tick labels, etc.


    ax.set_xticks(x)
    ax.set_xticklabels(x_labels)
    ax.legend(loc='lower right')
    ax.yaxis.set_major_formatter(mtick.PercentFormatter())
    # Axis styling.
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.spines['bottom'].set_color('#DDDDDD')
    ax.tick_params(bottom=False, left=False)
    ax.set_axisbelow(True)
    ax.yaxis.grid(True, color='#EEEEEE')
    ax.xaxis.grid(False)

    # Add axis and chart labels.
    ax.set_xlabel('# Solution')
    ax.set_ylabel('Relative performance')
    ax.set_title('Relative performance of potential solutions')

    #ax.bar_label(rects1, padding=3)
    #ax.bar_label(rects2, padding=3)
    #ax.bar_label(rects3, padding=3)
    fig.tight_layout()
    return fig

def create_evaluation_terminals(terminals, x_labels):
    labels = [i for i in range(1,len(terminals)+1)]
    x = np.arange(len(terminals))
    st.write(x)
    fig, ax = plt.subplots()
    ax.bar(x_labels,terminals)
    # Add some text for labels, title and custom x-axis tick labels, etc.


    ax.set_xticks(x)
    ax.set_xticklabels(x_labels)
    ax.legend(loc='lower right')
    ax.yaxis.set_major_formatter(mtick.PercentFormatter())
    # Axis styling.
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.spines['bottom'].set_color('#DDDDDD')
    ax.tick_params(bottom=False, left=False)
    ax.set_axisbelow(True)
    ax.yaxis.grid(True, color='#EEEEEE')
    ax.xaxis.grid(False)

    # Add axis and chart labels.
    ax.set_xlabel('# Solution')
    ax.set_ylabel('Relative performance')
    ax.set_title('Relative performance of potential solutions')

    #ax.bar_label(rects1, padding=3)
    #ax.bar_label(rects2, padding=3)
    #ax.bar_label(rects3, padding=3)
    fig.tight_layout()
    return fig


def create_relative_performance_final(makespan, tardiness, waiting, x_labels, header):
    labels = [i for i in range(1,len(makespan)+1)]
    x = np.arange(len(makespan))
    width = 0.25
    fig, ax = plt.subplots()
    rects1 = ax.bar(x - width , makespan*100, width, label='Makespan')
    rects2 = ax.bar(x, tardiness*100, width, label='Total Tardiness')
    rects3 = ax.bar(x + width , waiting*100, width, label='Total waiting time')
    # Add some text for labels, title and custom x-axis tick labels, etc.


    ax.set_xticks(x)
    ax.set_xticklabels(x_labels)
    ax.legend(loc='lower right')
    ax.yaxis.set_major_formatter(mtick.PercentFormatter())
    # Axis styling.
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.spines['bottom'].set_color('#DDDDDD')
    ax.tick_params(bottom=False, left=False)
    ax.set_axisbelow(True)
    ax.yaxis.grid(True, color='#EEEEEE')
    ax.xaxis.grid(False)

    # Add axis and chart labels.
    ax.set_xlabel('Dispatching rule')
    ax.set_ylabel('Relative performance')
    ax.set_title(header)

    #ax.bar_label(rects1, padding=3)
    #ax.bar_label(rects2, padding=3)
    #ax.bar_label(rects3, padding=3)
    fig.tight_layout()
    return fig

def create_schedule_visualization(results):
    schedule = pd.DataFrame(results)
    JOBS = sorted(list(schedule['Job'].unique()))
    MACHINES = sorted(list(schedule['Machine'].unique()))
    makespan = schedule['Finish'].max()

    bar_style = {'alpha': 1.0, 'lw': 25, 'solid_capstyle': 'butt'}
    text_style = {'color': 'white', 'weight': 'bold', 'ha': 'center', 'va': 'center'}
    colors = mpl.cm.Dark2.colors

    schedule.sort_values(by=['Job', 'Start'])
    schedule.set_index(['Job', 'Machine'], inplace=True)

    fig, ax = plt.subplots(2, 1, figsize=(20, 10 + (max(len(JOBS), len(MACHINES)) / 2)))

    for jdx, j in enumerate(JOBS, 1):
        for mdx, m in enumerate(MACHINES, 1):
            if (j, m) in schedule.index:
                xs = schedule.loc[(j, m), 'Start']
                xf = schedule.loc[(j, m), 'Finish']
                ax[0].plot([xs, xf], [jdx] * 2, c=colors[mdx % 7], **bar_style)
                ax[0].text((xs + xf) / 2, jdx, m, **text_style)
                ax[1].plot([xs, xf], [mdx] * 2, c=colors[jdx % 7], **bar_style)
                ax[1].text((xs + xf) / 2, mdx, j, **text_style)

    ax[0].set_title('Job Schedule')
    ax[0].set_ylabel('Job')
    ax[1].set_title('Machine Schedule')
    ax[1].set_ylabel('Machine')

    for idx, s in enumerate([JOBS, MACHINES]):
        ax[idx].set_ylim(0.5, len(s) + 0.5)
        ax[idx].set_yticks(range(1, 1 + len(s)))
        ax[idx].set_yticklabels(s)
        ax[idx].text(makespan, ax[idx].get_ylim()[0] - 0.2, "{0:0.1f}".format(makespan), ha='center', va='top')
        ax[idx].plot([makespan] * 2, ax[idx].get_ylim(), 'r--')
        ax[idx].set_xlabel('Time')
        ax[idx].grid(True)

    fig.tight_layout()
    return fig

def create_tree_visualization(best_solution):
    # create visualisation of tree of the final solution
    nodes, edges, labels = gp.graph(best_solution)

    st.write(nodes)
    st.write(edges)
    st.write(labels)

    fig, ax = plt.subplots(figsize=(14,10))
    g = nx.Graph()
    g.add_nodes_from(nodes)
    g.add_edges_from(edges)
    pos = nx.nx_agraph.graphviz_layout(g, prog='dot')
    nx.draw_networkx_nodes(g, pos)
    nx.draw_networkx_edges(g, pos)
    nx.draw_networkx_labels(g, pos, labels, font_size=4, font_color="whitesmoke")
    path = 'D:/PycharmProjects/05_DSS_hyper_heuristic/hyperheuristics/html_files'
    plt.savefig(f'{path}/nx_graph.png')
    image = Image.open(f'{path}/nx_graph.png')
    return image

def create_tree_visualization_final_solution(nodes, edges, labels):
    # create visualisation of tree of the final solution

    fig, ax = plt.subplots(figsize=(12,8), dpi=600)
    g = nx.Graph()
    g.add_nodes_from(nodes)
    g.add_edges_from(edges)
    pos = nx.nx_agraph.graphviz_layout(g, prog='dot')
    nx.draw_networkx_nodes(g, pos)
    nx.draw_networkx_edges(g, pos)
    nx.draw_networkx_labels(g, pos, labels, font_size=4, font_color="whitesmoke")
    path = 'D:/PycharmProjects/05_DSS_hyper_heuristic/hyperheuristics/html_files'
    plt.savefig(f'{path}/nx_graph.png')
    image = Image.open(f'{path}/nx_graph.png')
    return image

def create_radar_chart(project_name, solution1, solution2, objective_ideal, objective_nadir):
    # load data
    result_table = pd.read_excel(project_name + '.xlsx', sheet_name='ref_point', header=0,
                                 index_col=None)
    # prepare data
    makespan_solution1 = result_table[result_table['# Solution']==solution1]['Makespan']
    tardiness_solution1 = result_table[result_table['# Solution'] == solution1]['Total tardiness']
    waiting_solution1 = result_table[result_table['# Solution'] == solution1]['Total waiting time']

    makespan_solution2 = result_table[result_table['# Solution']==solution2]['Makespan']
    tardiness_solution2 = result_table[result_table['# Solution'] == solution2]['Total tardiness']
    waiting__solution2 = result_table[result_table['# Solution'] == solution2]['Total waiting time']

    makespan_solution1 = makespan_solution1.iloc[0]
    tardiness_solution1 = tardiness_solution1.iloc[0]
    waiting_solution1 = waiting_solution1.iloc[0]

    makespan_solution2 = makespan_solution2.iloc[0]
    tardiness_solution2 = tardiness_solution2.iloc[0]
    waiting__solution2 = waiting__solution2.iloc[0]

    def _scale_data(data, ranges):
        """scales data[1:] to ranges[0],
        inverts if the scale is reversed"""
        for d, (y1, y2) in zip(data[1:], ranges[1:]):
            assert (y1 <= d <= y2) or (y2 <= d <= y1)
        x1, x2 = ranges[0]
        d = data[0]
        # if x1 > x2:
        #    d = _invert(d, (x1, x2))
        #    x1, x2 = x2, x1
        sdata = [d]
        for d, (y1, y2) in zip(data[1:], ranges[1:]):
            # if y1 > y2:
            #    d = _invert(d, (y1, y2))
            #    y1, y2 = y2, y1
            sdata.append((d - y1) / (y2 - y1)
                             * (x2 - x1) + x1)
        return sdata


    class ComplexRadar():
        def __init__(self, fig, variables, ranges,
                     n_ordinate_levels=6):
            angles = np.arange(0, 360, 360. / len(variables))

            axes = [fig.add_axes([0.1, 0.1, 0.9, 0.9], polar=True,
                                 label="axes{}".format(i))
                    for i in range(len(variables))]
            l, text = axes[0].set_thetagrids(angles,
                                             labels=variables, size=12)
            [txt.set_rotation(angle - 90) for txt, angle
             in zip(text, angles)]
            for ax in axes[1:]:
                ax.patch.set_visible(False)
                ax.grid("off")
                ax.xaxis.set_visible(False)
            for i, ax in enumerate(axes):
                grid = np.linspace(*ranges[i],
                                   num=n_ordinate_levels)
                gridlabel = ["{}".format(round(x, 2))
                             for x in grid]
                # if ranges[i][0] > ranges[i][1]:
                # grid = grid[::-1] # hack to invert grid
                # gridlabels aren't reversed
                gridlabel[0] = ""  # clean up origin
                ax.set_rgrids(grid, labels=gridlabel, color='grey', size=7,
                              angle=angles[i])
                ax.spines["polar"].set_visible(True)
                ax.set_ylim(*ranges[i])
                ax.set_theta_offset(pi / 2)
                ax.set_theta_direction(-1)
            # variables for plotting
            self.angle = np.deg2rad(np.r_[angles, angles[0]])
            self.ranges = ranges
            self.ax = axes[0]

        def plot(self, data, *args, **kw):
            sdata = _scale_data(data, self.ranges)
            l = self.ax.plot(self.angle, np.r_[sdata, sdata[0]], *args, **kw)
            return l

        def fill(self, data, *args, **kw):
            sdata = _scale_data(data, self.ranges)
            self.ax.fill(self.angle, np.r_[sdata, sdata[0]], *args, **kw)

    # plotting
    fig1 = plt.figure(figsize=(6, 6))
    variables = ("Makespan", "Total tardiness", "Total waiting time")
    ranges = [(objective_nadir[0], objective_ideal[0]), (objective_nadir[1], objective_ideal[1]),
              (objective_nadir[2], objective_ideal[2])]
    radar = ComplexRadar(fig1, variables, ranges)
    lax = []
    index = [solution1, solution2]

    df = pd.DataFrame({
        "Makespan": pd.Series([makespan_solution1,
                                  makespan_solution2]),
        "Total tardiness": pd.Series([tardiness_solution1,
                                        tardiness_solution2]),
        "Total waiting time": pd.Series([waiting_solution1,
                                       waiting__solution2]),
    })

    for i, name in enumerate(index):
        data = df.iloc[i].values
        l, = radar.plot(data, label=name)
        lax.append(l)
        radar.fill(data, alpha=0.2)
    legendax = fig1.add_axes([0.8, 0.8, 0.1, .1])
    legendax.legend(handles=lax, labels=index, loc=3, bbox_to_anchor=(0, 0, 1, 1), bbox_transform=fig1.transFigure,
                    title='Solution')
    legendax.axis('off')
    return fig1

def save_values(project_name, ref_point, performance_name, performance_measure, opt_type, best_fitness, reference_point_input, best_solution, performance):
    wb = op.load_workbook(project_name + '.xlsx')
    ws_ideal_nadir =  wb["ideal_nadir"]
    ws_ref_point = wb["ref_point"]
    max_row = ws_ref_point.max_row
    if ref_point==None:
        st.header(performance_name[performance_measure])
        st.write(opt_type + " solution")
        st.write(best_fitness)
        if opt_type == 'best':
            ws_ideal_nadir['B'+str(performance_measure+2)] = best_fitness
        else:
            ws_ideal_nadir['C' + str(performance_measure + 2)] = best_fitness
    else:
        st.header('Best solution')
        st.write(best_fitness)
        ws_ref_point['A' + str(max_row + 1)] = str(reference_point_input)
        ws_ref_point['B' + str(max_row + 1)] = best_fitness
        ws_ref_point['C' + str(max_row + 1)] = str(best_solution)
        ws_ref_point['D' + str(max_row + 1)] = performance[0]
        ws_ref_point['E' + str(max_row + 1)] = performance[1]
        ws_ref_point['F' + str(max_row + 1)] = performance[2]
    wb.save(project_name+'.xlsx')
    wb.close()

def eliminate_solution(project_name, solution):
    wb = op.load_workbook(project_name + '.xlsx')
    ws_ref_point = wb["ref_point"]
    for row in ws_ref_point.rows:
        if row[0].value == solution:
            rownumber = row[0].row
    ws_ref_point['L'+str(rownumber)] = 1
    wb.save(project_name+'.xlsx')
    wb.close()

def div(left, right):
    try:
        return left / right
    except ZeroDivisionError:
        return 1

def ifte(condition, return_if_true, return_if_not_true):
    if condition >= 0:
        argument = return_if_true
    else:
        argument = return_if_not_true
    return argument

def load_test_problem(name):
    # TA15 benchmark problem
    number_machines, n_jobs = 15, 20
    data = """
     7 84  0 58 12 71  4 26  1 98  9 36  2 12 11 30 10 87 14 95  5 45  6 28 13 73  3 73  8 45 
     4 29  8 22  7 47  3 75  9 94 13 15 12  4  0 82 11 14 10 35  1 79  6 34  5 57 14 23  2 56 
     1 73  4 36  7 48 13 26  3 49  8 60 10 15  5 66 12 90 14 39  9  8  6 74  2 63  0 94 11 91 
     5  1 11 35  9 23 12 93  7 75  1 50  6 40 13 60  8 41  2  7  0 57 14 72  3 40  4 75 10  7 
     4 13 11 15 12 17  1 14  0 67  9 94  6 18 13 52  2 53 14 16  5 33 10 61  3 47  8 65  7 39 
     2 54  6 80  3 87  8 36 14 54  0 72  4 17 10 44 11 37  1 88  7 77 13 84 12 17  5 82  9 90 
     4  4 14 62  5 33 10 62  8 86  7 30  6 39  1 67  0 42 12 31  9 83 13 39 11 67  3 67  2 31 
     7 29 10 29 11 69 14 26  3 55  2 46  4 53  5 65  1 97 12 24  9 69  6 22 13 17  0 39  8 13 
    14 12 11 73  0 36 13 70  3 12  2 80  1 99  8 70  5 51  7 14  4 71 12 28  6 35 10 58  9 35 
     0 61  5 49 12 74  1 90 13 60 10 88  9  3  4 60  2 59  8 94 14 91 11 34  7 26  6  4  3 26 
     4 89  3 90  8 95 12 32  9 18 11 73  2  9 14 19  5 97  7 58 13 36  6 62 10 13  1 16  0  1 
     9 71  6 47  1 95  0  7 14 63  7 49 13 24 12 46  2 72 11 73  5 19  8 96 10 41  3 15  4 81 
     4 45  3  9  0 97 14 62 13 77  9 78  7 70  2 19 11 86  8 15 10 23  1 46  6 32 12  6  5 70 
    12 74 10 46  3 98  6  1  4 53  5 59  0 86  7 98  2 76  8 12 13 91 11 98 14 98  9 11  1 27 
    14 73  7 70  5 14  8 32 11 19  0 57  2 17 13 96 12 56  4 73  6 32  1  7 10 79  9 10  3 91 
     6 39 14 87 12 11  2 81  7  7  5 79  8 24 13  9 11 58  9 42  0 67  3 27  4 20  1 19 10 67 
     9 76  5 89 14 64 10 14 12 11  1 14  4 99 13 85  0 81 11  3  3 46  2 47  7 40  6 81  8 27 
     9 55 12 71  4  5 14 83 11 16  8  4  0 20  7 15  5 60  3  8  1 93 10 33  6 63 13 71  2 29 
    12 92  2 25  3  8 14 86  5 22  1 79  6 23 11 96 13 24  9 94  7 97 10 17  8 48  0 67  4 47 
     3  5 12 77 10 74  5 59 14 13  0 57  9 62  8 37 13 54  6 69 11 80  1 35  7 88  2 47  4 98 
         """
    TASKS = {}
    for job, line in enumerate(data.splitlines()[1:]):
        nums = line.split()
        prec = None
        for m, dur in zip(nums[::2], nums[1::2]):
            task = (job, int(m))
            TASKS[task] = {'dur': int(dur), 'prec': prec}
            prec = task
    due_date_list_jobs = []
    release_date_list = []
    total_processing_time = []
    due_date_tightness = 1.3
    for j in range(n_jobs):
        total_processing_time_current_job = 0
        for m in range(number_machines):
            total_processing_time_current_job += TASKS[j, m]['dur']
        total_processing_time.append(total_processing_time_current_job)
        due_date_list_jobs.append(due_date_tightness * total_processing_time_current_job)
        release_date_list.append(0)
    return TASKS, total_processing_time, due_date_list_jobs, release_date_list, n_jobs, number_machines

def load_test_problem_random(name):
    # random instance generator according to Taillard
    number_machines, n_jobs = 10, 10
    due_date_tightness = 1.3
    TASKS = {}
    due_date_list_jobs = []
    total_processing_time = []
    release_date_list = []
    sum_of_all_operations = 0
    for i in range(n_jobs):
        prec = None
        sum_proc_time = 0
        allowed_values = list(range(0, number_machines))
        for m in range(number_machines):
            dur = np.random.uniform(1, 99)
            sum_proc_time += dur
            machine = np.random.choice(allowed_values)
            task = (i, machine)
            TASKS[task] = {'dur': int(dur), 'prec': prec}
            prec = task
            allowed_values.remove(machine)
    for j in range(n_jobs):
        total_processing_time_current_job = 0
        for m in range(number_machines):
            total_processing_time_current_job += TASKS[j, m]['dur']
        total_processing_time.append(total_processing_time_current_job)
        sum_of_all_operations += total_processing_time_current_job
        due_date_list_jobs.append(due_date_tightness * total_processing_time_current_job)
        release_date_list.append(0)
    return TASKS, total_processing_time, due_date_list_jobs, release_date_list, n_jobs, number_machines, sum_of_all_operations

def load_test_problem_benchmark_TA(name):
    # TA15 benchmark problem
    number_machines, n_jobs = 15, 20
    data = """
         7 84  0 58 12 71  4 26  1 98  9 36  2 12 11 30 10 87 14 95  5 45  6 28 13 73  3 73  8 45 
         4 29  8 22  7 47  3 75  9 94 13 15 12  4  0 82 11 14 10 35  1 79  6 34  5 57 14 23  2 56 
         1 73  4 36  7 48 13 26  3 49  8 60 10 15  5 66 12 90 14 39  9  8  6 74  2 63  0 94 11 91 
         5  1 11 35  9 23 12 93  7 75  1 50  6 40 13 60  8 41  2  7  0 57 14 72  3 40  4 75 10  7 
         4 13 11 15 12 17  1 14  0 67  9 94  6 18 13 52  2 53 14 16  5 33 10 61  3 47  8 65  7 39 
         2 54  6 80  3 87  8 36 14 54  0 72  4 17 10 44 11 37  1 88  7 77 13 84 12 17  5 82  9 90 
         4  4 14 62  5 33 10 62  8 86  7 30  6 39  1 67  0 42 12 31  9 83 13 39 11 67  3 67  2 31 
         7 29 10 29 11 69 14 26  3 55  2 46  4 53  5 65  1 97 12 24  9 69  6 22 13 17  0 39  8 13 
        14 12 11 73  0 36 13 70  3 12  2 80  1 99  8 70  5 51  7 14  4 71 12 28  6 35 10 58  9 35 
         0 61  5 49 12 74  1 90 13 60 10 88  9  3  4 60  2 59  8 94 14 91 11 34  7 26  6  4  3 26 
         4 89  3 90  8 95 12 32  9 18 11 73  2  9 14 19  5 97  7 58 13 36  6 62 10 13  1 16  0  1 
         9 71  6 47  1 95  0  7 14 63  7 49 13 24 12 46  2 72 11 73  5 19  8 96 10 41  3 15  4 81 
         4 45  3  9  0 97 14 62 13 77  9 78  7 70  2 19 11 86  8 15 10 23  1 46  6 32 12  6  5 70 
        12 74 10 46  3 98  6  1  4 53  5 59  0 86  7 98  2 76  8 12 13 91 11 98 14 98  9 11  1 27 
        14 73  7 70  5 14  8 32 11 19  0 57  2 17 13 96 12 56  4 73  6 32  1  7 10 79  9 10  3 91 
         6 39 14 87 12 11  2 81  7  7  5 79  8 24 13  9 11 58  9 42  0 67  3 27  4 20  1 19 10 67 
         9 76  5 89 14 64 10 14 12 11  1 14  4 99 13 85  0 81 11  3  3 46  2 47  7 40  6 81  8 27 
         9 55 12 71  4  5 14 83 11 16  8  4  0 20  7 15  5 60  3  8  1 93 10 33  6 63 13 71  2 29 
        12 92  2 25  3  8 14 86  5 22  1 79  6 23 11 96 13 24  9 94  7 97 10 17  8 48  0 67  4 47 
         3  5 12 77 10 74  5 59 14 13  0 57  9 62  8 37 13 54  6 69 11 80  1 35  7 88  2 47  4 98 
             """
    TASKS = {}
    for job, line in enumerate(data.splitlines()[1:]):
        nums = line.split()
        prec = None
        for m, dur in zip(nums[::2], nums[1::2]):
            task = (job, int(m))
            TASKS[task] = {'dur': int(dur), 'prec': prec}
            prec = task
    due_date_list_jobs = []
    release_date_list = []
    total_processing_time = []
    due_date_tightness = 1.3
    for j in range(n_jobs):
        total_processing_time_current_job = 0
        for m in range(number_machines):
            total_processing_time_current_job += TASKS[j, m]['dur']
        total_processing_time.append(total_processing_time_current_job)
        due_date_list_jobs.append(due_date_tightness * total_processing_time_current_job)
        release_date_list.append(0)
    return TASKS, total_processing_time, due_date_list_jobs, release_date_list, n_jobs, number_machines

def geneticprogamming(performance_measure, ref_point, opt_type, project_name, reference_point_input, nb_generations, population_size, crossover_probability, mutation_probability, max_depth_crossover, max_depth_mutation, nb_simulations):
    performance_name = ["makespan", "number tardy jobs", "total tardiness"]
    def div(left, right):
        try:
            return left / right
        except ZeroDivisionError:
            return 1

    pset = gp.PrimitiveSet("MAIN", 4)
    pset.addPrimitive(operator.add, 2)
    pset.addPrimitive(operator.sub, 2)
    pset.addPrimitive(operator.mul, 2)
    pset.addPrimitive(min, 2)
    pset.addPrimitive(max, 2)
    pset.addPrimitive(div, 2)
    #pset.addPrimitive(operator.neg, 1)
    #pset.addPrimitive(math.cos, 1)
    #pset.addPrimitive(math.sin, 1)
    #randomname = str(rd.randint(1, 100000000000000)) + str(rd.randint(1, 100000000000000)) + str(rd.randint(1, 100000000000000))
    # Problem bei Konstante da jedes mal neuer Name generiert werden muss
    #try:
    #    pset.addEphemeralConstant(randomname, lambda: rd.uniform(-1, 1))
    #except:
    #    randomname = str(rd.randint(1, 100000000000000)) + str(rd.randint(1, 100000000000000)) + str(
    #        rd.randint(1, 100000000000000))
    #    pset.addEphemeralConstant(randomname, lambda: rd.uniform(-1, 1))
    #maybe change here to random between 0 and 10 as mentioned in overleaf
    pset.renameArguments(ARG0='PT')
    pset.renameArguments(ARG1='SPT')
    pset.renameArguments(ARG2='RPT')
    pset.renameArguments(ARG3='DD')

    if opt_type=="best":
        creator.create("FitnessMin", base.Fitness, weights=(-1.0,))
        creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMin)
    else:
        creator.create("FitnessMax", base.Fitness, weights=(1.0,))
        creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMax)


    toolbox = base.Toolbox()
    toolbox.register("expr", gp.genHalfAndHalf, pset=pset, min_=2, max_=6)
    toolbox.register("individual", tools.initIterate, creator.Individual, toolbox.expr)
    toolbox.register("population", tools.initRepeat, list, toolbox.individual)
    toolbox.register("compile", gp.compile, pset=pset)



    def evalSymbReg(individual):
        func = toolbox.compile(expr=individual)
        def visualize(results):
            schedule = pd.DataFrame(results)
            JOBS = sorted(list(schedule['Job'].unique()))
            MACHINES = sorted(list(schedule['Machine'].unique()))
            makespan = schedule['Finish'].max()

            bar_style = {'alpha': 1.0, 'lw': 25, 'solid_capstyle': 'butt'}
            text_style = {'color': 'white', 'weight': 'bold', 'ha': 'center', 'va': 'center'}
            colors = mpl.cm.Dark2.colors

            schedule.sort_values(by=['Job', 'Start'])
            schedule.set_index(['Job', 'Machine'], inplace=True)

            fig, ax = plt.subplots(2, 1, figsize=(12, 5 + (len(JOBS) + len(MACHINES)) / 4))

            for jdx, j in enumerate(JOBS, 1):
                for mdx, m in enumerate(MACHINES, 1):
                    if (j, m) in schedule.index:
                        xs = schedule.loc[(j, m), 'Start']
                        xf = schedule.loc[(j, m), 'Finish']
                        ax[0].plot([xs, xf], [jdx] * 2, c=colors[mdx % 7], **bar_style)
                        ax[0].text((xs + xf) / 2, jdx, m, **text_style)
                        ax[1].plot([xs, xf], [mdx] * 2, c=colors[jdx % 7], **bar_style)
                        ax[1].text((xs + xf) / 2, mdx, j, **text_style)

            ax[0].set_title('Job Schedule')
            ax[0].set_ylabel('Job')
            ax[1].set_title('Machine Schedule')
            ax[1].set_ylabel('Machine')

            for idx, s in enumerate([JOBS, MACHINES]):
                ax[idx].set_ylim(0.5, len(s) + 0.5)
                ax[idx].set_yticks(range(1, 1 + len(s)))
                ax[idx].set_yticklabels(s)
                ax[idx].text(makespan, ax[idx].get_ylim()[0] - 0.2, "{0:0.1f}".format(makespan), ha='center', va='top')
                ax[idx].plot([makespan] * 2, ax[idx].get_ylim(), 'r--')
                ax[idx].set_xlabel('Time')
                ax[idx].grid(True)

            fig.tight_layout()
            plt.show()

        #def source(env, number, machine):
        #    """Source generates jobs randomly"""
        #    for i in range(number):
        #        processing_time = np.random.uniform(avg_processing_time - delta_processing_time,
        #                                            avg_processing_time + delta_processing_time)
        #        if i == 0:
        #            release_time = np.random.uniform(avg_job_interarrival_time - delta_release_date,
        #                                             avg_job_interarrival_time + delta_release_date)
        #        else:
        #            release_time = release_time_list[i - 1] + np.random.uniform(
        #                avg_job_interarrival_time - delta_release_date,
        #                avg_job_interarrival_time + delta_release_date)

        #        due_date = release_time + k * processing_time
        #        due_date_list.append(due_date)
        #        processing_time_list.append(processing_time)
        #        release_time_list.append(release_time)
        #        c = job(env, f'Job {i + 1}', machine, processing_time=processing_time,
        #                total_processing_time=total_processing_time)
        #        env.process(c)
        #        t = 0
        #        yield env.timeout(t)

        def job(env, name, machine, processing_time, total_processing_time, remaining_processing_time, due_date,
                release_date):
            """Job arrives, is served and leaves."""
            # arrive = release_time
            # print('%7.4f %s: Arrived' % (arrive, name))

            with machine.request(priority=func(processing_time, total_processing_time, remaining_processing_time, due_date)) as req:
                yield req
                # wait = env.now - arrive # waiting time of job
                job_start = env.now

                # We got to the counter
                # print('%7.4f %s: Waited %6.3f' % (env.now, name, wait))

                yield env.timeout(processing_time)
                # print('%7.4f %s: Finished' % (env.now, name))

                # Flow Time
                # flow_time.append(env.now-arrive)
                remaining_processing_time -= processing_time
                job_prec = name
                machine_prec = machine_list.index(machine)
                schedule.append([job_prec, machine_prec])
                #results.append({'Job': f'J{job_prec}',
                #                'Machine': f'M{machine_prec}',
                #                'Start': job_start,
                #                'Duration': processing_time,
                #                'Finish': env.now})
                job_finished = 'Yes'
                for j in range(n_jobs):
                    for m in range(number_machines):
                        if TASKS[j, m]['prec'] == (job_prec, machine_prec):
                            machine = machine_list[m]
                            processing_time = TASKS[j, m]['dur']
                            job_finished = 'No'
                            env.process(
                                job(env, j, machine, processing_time, total_processing_time, remaining_processing_time,
                                    due_date, release_date))
                if job_finished == 'Yes':
                    # Completion time
                    completion_time.append(env.now)
                    # Tardiness of job
                    tardiness.append(max(env.now - due_date, 0))
                    # Tardy jobs
                    if max(env.now - due_date, 0) > 0:
                        tardy_jobs.append(1)
                    else:
                        tardy_jobs.append(0)

        number_simulations = nb_simulations
        avg_makespan_list = []
        avg_mean_tardiness_list = []
        avg_max_tardiness_list = []
        avg_total_tardiness_list = []
        avg_number_tardy_jobs_list = []

        for simulations in range(number_simulations):
            #np.random.seed(10)
            # random job generator
            #number_machines = 10
            #n_jobs = 10
            #duedate_tightness = 1.5
            #TASKS = {}
            #release_date_list = []
            #due_date_list_jobs = []
            #total_processing_time = []
            #for i in range(n_jobs):
            #    prec = None
            #    release_time = np.random.uniform(0, 40)
            #    sum_proc_time = 0
            #    allowed_values = list(range(0, 10))
            #    for m in range(number_machines):
            #        dur = np.random.uniform(number_machines / 2, number_machines * 2)
            #        sum_proc_time += dur
            #        machine = np.random.choice(allowed_values)
            #        task = (i, machine)
            #        TASKS[task] = {'dur': float(dur), 'prec': prec}
            #        prec = task
            #        allowed_values.remove(machine)
            #    due_date = release_time + duedate_tightness * sum_proc_time
            #    release_date_list.append(release_time)
            #    due_date_list_jobs.append(due_date)

            #for j in range(n_jobs):
            #    total_processing_time_current_job = 0
            #    for m in range(number_machines):
            #        total_processing_time_current_job += TASKS[j, m]['dur']
            #    total_processing_time.append(total_processing_time_current_job)

            TASKS, total_processing_time, due_date_list_jobs, release_date_list, n_jobs, number_machines = load_test_problem_random(
                name='random')

            # print(TASKS['J0', 'M0']['prec'])
            # print(pd.DataFrame(TASKS).T)

            #number_simulations = 1
            #R_processing_time = 0.4
            #avg_processing_time = 10
            #processing_time = []
            #duedate_tightness = 2
            #duedate_variability = 0.3
            #machine_utilization = 0.7
            #job_interarrival_tightness = 1
            schedule = []
            #release_time = []
            tardiness = []
            tardy_jobs = []
            completion_time = []
            #flow_time = []
            #due_date_list = []
            #release_time_list = []
            #processing_time_list = []
            #results = []
            #remaining_processing_time_current_job = []

            #n_jobs_original = n_jobs
            #delta_processing_time = R_processing_time * avg_processing_time
            #avg_job_interarrival_time = avg_processing_time / machine_utilization
            #delta_release_date = job_interarrival_tightness * avg_job_interarrival_time
            #release_time.append(np.random.uniform(avg_job_interarrival_time - delta_release_date,
            #                                      avg_job_interarrival_time + delta_release_date))
            #delta_duedate = duedate_tightness * duedate_variability
            #k = np.random.uniform(duedate_tightness - delta_duedate, duedate_tightness + delta_duedate)

            env = simpy.Environment()

            # Start processes and run
            machine_0 = simpy.PriorityResource(env, capacity=1)
            machine_1 = simpy.PriorityResource(env, capacity=1)
            machine_2 = simpy.PriorityResource(env, capacity=1)
            machine_3 = simpy.PriorityResource(env, capacity=1)
            machine_4 = simpy.PriorityResource(env, capacity=1)
            machine_5 = simpy.PriorityResource(env, capacity=1)
            machine_6 = simpy.PriorityResource(env, capacity=1)
            machine_7 = simpy.PriorityResource(env, capacity=1)
            machine_8 = simpy.PriorityResource(env, capacity=1)
            machine_9 = simpy.PriorityResource(env, capacity=1)
            machine_10 = simpy.PriorityResource(env, capacity=1)
            machine_11 = simpy.PriorityResource(env, capacity=1)
            machine_12 = simpy.PriorityResource(env, capacity=1)
            machine_13 = simpy.PriorityResource(env, capacity=1)
            machine_14 = simpy.PriorityResource(env, capacity=1)
            # machine = simpy.Resource(env, capacity=1)

            machine_list = [machine_0, machine_1, machine_2, machine_3, machine_4, machine_5, machine_6, machine_7,
                            machine_8, machine_9, machine_10, machine_11, machine_12, machine_13, machine_14]

            for j in range(n_jobs):
                for m in range(number_machines):
                    if TASKS[j, m]['prec'] == None:
                        current_machine = machine_list[m]
                        processing_time_new = float(TASKS[j, m]['dur'])
                        job_new = j
                        total_processing_time_current_job = total_processing_time[j]
                        remaining_processing_time_current_job = total_processing_time_current_job
                        release_date_current_job = release_date_list[j]
                        due_date_current_job = due_date_list_jobs[j]
                        env.process(job(env, job_new, current_machine, processing_time=processing_time_new,
                                        total_processing_time=total_processing_time_current_job,
                                        remaining_processing_time=remaining_processing_time_current_job,
                                        due_date=due_date_current_job, release_date=release_date_current_job))

            env.run()

            # for i in range(number_simulations):
            #    env.process(source(env, n_jobs, machine_1))
            #    env.run()

            # Post processing
            # calculate and performance measures of the current simulation run
            total_tardiness = sum(tardiness)
            mean_tardiness = mean(tardiness)
            max_tardiness = max(tardiness)
            number_tardy_jobs = sum(tardy_jobs)
            makespan = max(completion_time)
            # mean_flow_time = mean(flow_time)
            # max_flow_time = max(flow_time)
            # print('Release Time')
            # print(release_time_list)
            # print('processing time')
            # print(processing_time_list)
            # print('Due Dates')
            # print(due_date_list)
            # print(f'Total Tardiness: {total_tardiness}')
            # print(f'Mean Tardiness: {mean_tardiness}')
            # print(f'Max Tardiness: {max_tardiness}')
            # print(f'Number of tardy Jobs: {number_tardy_jobs}')
            # print(completion_time)
            # print(f'Makespan: {makespan}')
            # print(f'Mean flow time: {mean_flow_time}')
            # print(f'Max flow time: {max_flow_time}')
            # print(results)

            # add performance measures of current simulation run to the list for all runs
            avg_makespan_list.append(makespan)
            avg_mean_tardiness_list.append(mean_tardiness)
            avg_max_tardiness_list.append(max_tardiness)
            avg_total_tardiness_list.append(total_tardiness)
            avg_number_tardy_jobs_list.append(number_tardy_jobs)

            # visualize(results)

        # calculate and print the performance measures after all simulation runs
        avg_makespan = mean(avg_makespan_list)
        avg_mean_tardiness = mean(avg_mean_tardiness_list)
        avg_max_tardiness = mean(avg_max_tardiness_list)
        avg_total_tardiness = mean(avg_total_tardiness_list)
        avg_number_tardy_jobs = mean(avg_number_tardy_jobs_list)
        performance = [avg_makespan, avg_number_tardy_jobs, avg_total_tardiness]
        lambda_performance = [0.33, 0.33, 0.33]
        reference_point = reference_point_input
        roh = 0.0001
        # calculate the fitness value according to the achievement scalarizing function
        if ref_point==None:
            fitness = performance[performance_measure]
        else:
            fitness = max(lambda_performance[i]*(performance[i]-reference_point[i]) for i in range(0,len(reference_point))) + roh * sum(lambda_performance[i]*(performance[i]-reference_point[i]) for i in range(0,len(reference_point)))
        return fitness,


    toolbox.register("evaluate", evalSymbReg)
    toolbox.register("select", tools.selTournament, tournsize=5)
    toolbox.register("mate", gp.cxOnePoint)
    toolbox.register("expr_mut", gp.genFull, min_=0, max_=2)
    toolbox.register("mutate", gp.mutUniform, expr=toolbox.expr_mut, pset=pset)

    toolbox.decorate("mate", gp.staticLimit(key=operator.attrgetter("height"), max_value=max_depth_crossover))
    toolbox.decorate("mutate", gp.staticLimit(key=operator.attrgetter("height"), max_value=max_depth_mutation))

    #random.seed(318)

    pop = toolbox.population(n=population_size)
    hof = tools.HallOfFame(1)

    stats_fit = tools.Statistics(lambda ind: ind.fitness.values)
    stats_size = tools.Statistics(len)
    mstats = tools.MultiStatistics(fitness=stats_fit, size=stats_size)
    mstats.register("avg", np.mean)
    mstats.register("std", np.std)
    mstats.register("min", np.min)
    mstats.register("max", np.max)

    pop, log = algorithms.eaSimple(pop, toolbox, crossover_probability, mutation_probability, nb_generations, stats=mstats,
                                   halloffame=hof, verbose=True)

    # extract statistics:
    avgFitnessValues  = log.chapters['fitness'].select("avg")
    minFitnessValues = log.chapters['fitness'].select("min")
    maxFitnessValues = log.chapters['fitness'].select("max")
    stdFitnessValues = log.chapters['fitness'].select("std")
    nb_generation = log.select("gen")
    nevals = log.select('nevals')

    # plot statistics:
    # sns.set_style("whitegrid")
    fig, ax = plt.subplots()
    #ax.boxplot(minFitnessValues)
    mins = np.array(minFitnessValues)
    maxes = np.array(maxFitnessValues)
    means = np.array(avgFitnessValues)
    std = np.array(stdFitnessValues)
    gen = np.array(nb_generation)
    st.write(mins, maxes, means, std, gen)

    # create stacked errorbars:

    #ax.scatter(nb_generation, avgFitnessValues)
    ax.errorbar(gen, means, std, fmt='ok', lw=3)
    ax.errorbar(gen, means, [means - mins, maxes - means],
                 fmt='.k', ecolor='gray', lw=1)
    ax.set_xlabel('Generation')
    ax.set_ylabel('Average Fitness')
    st.header('Evolution process')
    st.pyplot(fig)

    #print("best solution:")
    #print(hof[0])
    best = hof.items[0]
    best_fitness = best.fitness.values[0]
    st.header('Dispatching rule of the best solution')
    st.write(str(best))

    # create visualisation of tree of the final solution
    nodes, edges, labels = gp.graph(hof.items[0])
    fig, ax = plt.subplots()
    g = nx.Graph()
    g.add_nodes_from(nodes)
    g.add_edges_from(edges)
    pos = nx.nx_agraph.graphviz_layout(g, prog='dot')
    nx.draw_networkx_nodes(g, pos)
    nx.draw_networkx_edges(g, pos)
    nx.draw_networkx_labels(g, pos, labels, font_size=6, font_color="whitesmoke")
    path = 'D:/PycharmProjects/05_DSS_hyper_heuristic/hyperheuristics/html_files'
    plt.savefig(f'{path}/nx_graph.png')
    image = Image.open(f'{path}/nx_graph.png')
    st.header('Primitive tree of the best solution')
    st.image(image,  use_column_width=True)

    wb = op.load_workbook(project_name + '.xlsx')
    ws_ideal_nadir =  wb["ideal_nadir"]
    ws_ref_point = wb["ref_point"]
    max_row = ws_ref_point.max_row
    if ref_point==None:
        st.header(performance_name[performance_measure])
        st.write(opt_type + " solution")
        st.write(best_fitness)
        if opt_type == 'best':
            ws_ideal_nadir['B'+str(performance_measure+2)] = best_fitness
        else:
            ws_ideal_nadir['C' + str(performance_measure + 2)] = best_fitness
    else:
        st.header('Best solution')
        st.write(best_fitness)
        ws_ref_point['A' + str(max_row + 1)] = str(reference_point_input)
        #ws_ref_point['A'+str(max_row+1)] = "["+str(reference_point_input[0]) +", " + str(reference_point_input[1]) +", "+ str(reference_point_input[2]) +"]"
        ws_ref_point['B' + str(max_row + 1)] = best_fitness
        ws_ref_point['C' + str(max_row + 1)] = str(best)

    print('best fitness slow \t\t', best_fitness)

    wb.save(project_name+'.xlsx')
    wb.close()
    #st.write(pop)
    #st.write(log)
    #st.stop()

def geneticprogamming_speedup(performance_measure, ref_point, opt_type, project_name, reference_point_input, nb_generations, population_size, crossover_probability, mutation_probability, max_depth_crossover, max_depth_mutation, nb_simulations):
    performance_name = ["makespan", "number tardy jobs", "total tardiness"]
    def div(left, right):
        try:
            return left / right
        except ZeroDivisionError:
            return 1

    pset = gp.PrimitiveSet("MAIN", 4)
    pset.addPrimitive(operator.add, 2)
    pset.addPrimitive(operator.sub, 2)
    pset.addPrimitive(operator.mul, 2)
    pset.addPrimitive(min, 2)
    pset.addPrimitive(max, 2)
    pset.addPrimitive(div, 2)
    #pset.addPrimitive(operator.neg, 1)
    #pset.addPrimitive(math.cos, 1)
    #pset.addPrimitive(math.sin, 1)
    #randomname = str(rd.randint(1, 100000000000000)) + str(rd.randint(1, 100000000000000)) + str(rd.randint(1, 100000000000000))
    # Problem bei Konstante da jedes mal neuer Name generiert werden muss
    #try:
    #    pset.addEphemeralConstant(randomname, lambda: rd.uniform(-1, 1))
    #except:
    #    randomname = str(rd.randint(1, 100000000000000)) + str(rd.randint(1, 100000000000000)) + str(
    #        rd.randint(1, 100000000000000))
    #    pset.addEphemeralConstant(randomname, lambda: rd.uniform(-1, 1))
    #maybe change here to random between 0 and 10 as mentioned in overleaf
    pset.renameArguments(ARG0='PT')
    pset.renameArguments(ARG1='SPT')
    pset.renameArguments(ARG2='RPT')
    pset.renameArguments(ARG3='DD')

    if opt_type=="best":
        creator.create("FitnessMin", base.Fitness, weights=(-1.0,))
        creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMin)
    else:
        creator.create("FitnessMax", base.Fitness, weights=(1.0,))
        creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMax)



    toolbox = base.Toolbox()
    toolbox.register("expr", gp.genHalfAndHalf, pset=pset, min_=2, max_=6)
    toolbox.register("individual", tools.initIterate, creator.Individual, toolbox.expr)
    toolbox.register("population", tools.initRepeat, list, toolbox.individual)
    toolbox.register("compile", gp.compile, pset=pset)


    # random job generator
    np.random.seed(10)
    number_machines, n_jobs, duedate_tightness = 10, 100, 1.5
    TASKS = {}
    release_date_list, due_date_list_jobs, total_processing_time = [], [], []
    for i in range(n_jobs):
        prec = None
        release_time = np.random.uniform(0, 40)
        sum_proc_time = 0
        allowed_values = list(range(0, 10))
        for m in range(number_machines):
            dur = np.random.uniform(number_machines / 2, number_machines * 2)
            sum_proc_time += dur
            machine = np.random.choice(allowed_values)
            task = (i, machine)
            TASKS[task] = {'dur': float(dur), 'prec': prec}
            prec = task
            allowed_values.remove(machine)
        due_date = release_time + duedate_tightness * sum_proc_time
        release_date_list.append(release_time)
        due_date_list_jobs.append(due_date)
        total_processing_time.append(sum_proc_time)


    def evalSymbReg(individual):
        func = toolbox.compile(expr=individual)

        def visualize(results):
            schedule = pd.DataFrame(results)
            JOBS = sorted(list(schedule['Job'].unique()))
            MACHINES = sorted(list(schedule['Machine'].unique()))
            makespan = schedule['Finish'].max()

            bar_style = {'alpha': 1.0, 'lw': 25, 'solid_capstyle': 'butt'}
            text_style = {'color': 'white', 'weight': 'bold', 'ha': 'center', 'va': 'center'}
            colors = mpl.cm.Dark2.colors

            schedule.sort_values(by=['Job', 'Start'])
            schedule.set_index(['Job', 'Machine'], inplace=True)

            fig, ax = plt.subplots(2, 1, figsize=(12, 5 + (len(JOBS) + len(MACHINES)) / 4))

            for jdx, j in enumerate(JOBS, 1):
                for mdx, m in enumerate(MACHINES, 1):
                    if (j, m) in schedule.index:
                        xs = schedule.loc[(j, m), 'Start']
                        xf = schedule.loc[(j, m), 'Finish']
                        ax[0].plot([xs, xf], [jdx] * 2, c=colors[mdx % 7], **bar_style)
                        ax[0].text((xs + xf) / 2, jdx, m, **text_style)
                        ax[1].plot([xs, xf], [mdx] * 2, c=colors[jdx % 7], **bar_style)
                        ax[1].text((xs + xf) / 2, mdx, j, **text_style)

            ax[0].set_title('Job Schedule')
            ax[0].set_ylabel('Job')
            ax[1].set_title('Machine Schedule')
            ax[1].set_ylabel('Machine')

            for idx, s in enumerate([JOBS, MACHINES]):
                ax[idx].set_ylim(0.5, len(s) + 0.5)
                ax[idx].set_yticks(range(1, 1 + len(s)))
                ax[idx].set_yticklabels(s)
                ax[idx].text(makespan, ax[idx].get_ylim()[0] - 0.2, "{0:0.1f}".format(makespan), ha='center', va='top')
                ax[idx].plot([makespan] * 2, ax[idx].get_ylim(), 'r--')
                ax[idx].set_xlabel('Time')
                ax[idx].grid(True)

            fig.tight_layout()
            plt.show()

        def source(env, number, machine):
            """Source generates jobs randomly"""
            for i in range(number):
                processing_time = np.random.uniform(avg_processing_time - delta_processing_time,
                                                    avg_processing_time + delta_processing_time)
                if i == 0:
                    release_time = np.random.uniform(avg_job_interarrival_time - delta_release_date,
                                                     avg_job_interarrival_time + delta_release_date)
                else:
                    release_time = release_time_list[i - 1] + np.random.uniform(
                        avg_job_interarrival_time - delta_release_date,
                        avg_job_interarrival_time + delta_release_date)

                due_date = release_time + k * processing_time
                due_date_list.append(due_date)
                processing_time_list.append(processing_time)
                release_time_list.append(release_time)
                c = job(env, f'Job {i + 1}', machine, processing_time=processing_time,
                        total_processing_time=total_processing_time)
                env.process(c)
                t = 0
                yield env.timeout(t)

        def job(env, name, machine, processing_time, total_processing_time, remaining_processing_time, due_date,
                release_date):
            """Job arrives, is served and leaves."""
            # arrive = release_time
            # print('%7.4f %s: Arrived' % (arrive, name))

            with machine.request(priority=func(processing_time, total_processing_time, remaining_processing_time, due_date)) as req:
                yield req
                # wait = env.now - arrive # waiting time of job
                job_start = env.now

                # We got to the counter
                # print('%7.4f %s: Waited %6.3f' % (env.now, name, wait))

                yield env.timeout(processing_time)
                # print('%7.4f %s: Finished' % (env.now, name))

                # Flow Time
                # flow_time.append(env.now-arrive)
                remaining_processing_time -= processing_time
                job_prec = name
                machine_prec = machine_list.index(machine)
                schedule.append([job_prec, machine_prec])
                #results.append({'Job': f'J{job_prec}',
                #                'Machine': f'M{machine_prec}',
                #                'Start': job_start,
                #                'Duration': processing_time,
                #                'Finish': env.now})
                job_finished = 'Yes'
                for j in range(10):
                    for m in range(10):
                        if TASKS[j, m]['prec'] == (job_prec, machine_prec):
                            machine = machine_list[m]
                            processing_time = TASKS[j, m]['dur']
                            job_finished = 'No'
                            env.process(
                                job(env, j, machine, processing_time, total_processing_time, remaining_processing_time,
                                    due_date, release_date))
                if job_finished == 'Yes':
                    # Completion time
                    completion_time.append(env.now)
                    # Tardiness of job
                    tardiness.append(max(env.now - due_date, 0))
                    # Tardy jobs
                    if max(env.now - due_date, 0) > 0:
                        tardy_jobs.append(1)
                    else:
                        tardy_jobs.append(0)

        avg_makespan_list = []
        avg_mean_tardiness_list = []
        avg_max_tardiness_list = []
        avg_total_tardiness_list = []
        avg_number_tardy_jobs_list = []

        for simulations in range(nb_simulations):




            # print(TASKS['J0', 'M0']['prec'])
            # print(pd.DataFrame(TASKS).T)

            number_simulations = 1
            R_processing_time = 0.4
            avg_processing_time = 10
            processing_time = []
            duedate_tightness = 2
            duedate_variability = 0.3
            machine_utilization = 0.7
            job_interarrival_tightness = 1
            schedule = []
            release_time = []
            tardiness = []
            tardy_jobs = []
            completion_time = []
            flow_time = []
            due_date_list = []
            release_time_list = []
            processing_time_list = []
            results = []
            remaining_processing_time_current_job = []

            n_jobs_original = n_jobs
            delta_processing_time = R_processing_time * avg_processing_time
            avg_job_interarrival_time = avg_processing_time / machine_utilization
            delta_release_date = job_interarrival_tightness * avg_job_interarrival_time
            release_time.append(np.random.uniform(avg_job_interarrival_time - delta_release_date,
                                                  avg_job_interarrival_time + delta_release_date))
            delta_duedate = duedate_tightness * duedate_variability
            k = np.random.uniform(duedate_tightness - delta_duedate, duedate_tightness + delta_duedate)

            env = simpy.Environment()

            # Start processes and run
            machine_0 = simpy.PriorityResource(env, capacity=1)
            machine_1 = simpy.PriorityResource(env, capacity=1)
            machine_2 = simpy.PriorityResource(env, capacity=1)
            machine_3 = simpy.PriorityResource(env, capacity=1)
            machine_4 = simpy.PriorityResource(env, capacity=1)
            machine_5 = simpy.PriorityResource(env, capacity=1)
            machine_6 = simpy.PriorityResource(env, capacity=1)
            machine_7 = simpy.PriorityResource(env, capacity=1)
            machine_8 = simpy.PriorityResource(env, capacity=1)
            machine_9 = simpy.PriorityResource(env, capacity=1)
            # machine = simpy.Resource(env, capacity=1)

            machine_list = [machine_0, machine_1, machine_2, machine_3, machine_4, machine_5, machine_6, machine_7,
                            machine_8, machine_9]

            for j in range(n_jobs):
                for m in range(number_machines):
                    if TASKS[j, m]['prec'] == None:
                        current_machine = machine_list[m]
                        processing_time_new = float(TASKS[j, m]['dur'])
                        job_new = j
                        total_processing_time_current_job = total_processing_time[j]
                        remaining_processing_time_current_job = total_processing_time_current_job
                        release_date_current_job = release_date_list[j]
                        due_date_current_job = due_date_list_jobs[j]
                        env.process(job(env, job_new, current_machine, processing_time=processing_time_new,
                                        total_processing_time=total_processing_time_current_job,
                                        remaining_processing_time=remaining_processing_time_current_job,
                                        due_date=due_date_current_job, release_date=release_date_current_job))

            env.run()

            # for i in range(number_simulations):
            #    env.process(source(env, n_jobs, machine_1))
            #    env.run()

            # Post processing
            # calculate and performance measures of the current simulation run
            total_tardiness = sum(tardiness)
            mean_tardiness = mean(tardiness)
            max_tardiness = max(tardiness)
            number_tardy_jobs = sum(tardy_jobs)
            makespan = max(completion_time)
            # mean_flow_time = mean(flow_time)
            # max_flow_time = max(flow_time)
            # print('Release Time')
            # print(release_time_list)
            # print('processing time')
            # print(processing_time_list)
            # print('Due Dates')
            # print(due_date_list)
            # print(f'Total Tardiness: {total_tardiness}')
            # print(f'Mean Tardiness: {mean_tardiness}')
            # print(f'Max Tardiness: {max_tardiness}')
            # print(f'Number of tardy Jobs: {number_tardy_jobs}')
            # print(completion_time)
            # print(f'Makespan: {makespan}')
            # print(f'Mean flow time: {mean_flow_time}')
            # print(f'Max flow time: {max_flow_time}')
            # print(results)

            # add performance measures of current simulation run to the list for all runs
            avg_makespan_list.append(makespan)
            avg_mean_tardiness_list.append(mean_tardiness)
            avg_max_tardiness_list.append(max_tardiness)
            avg_total_tardiness_list.append(total_tardiness)
            avg_number_tardy_jobs_list.append(number_tardy_jobs)

            # visualize(results)

        # calculate and print the performance measures after all simulation runs
        avg_makespan = mean(avg_makespan_list)
        avg_mean_tardiness = mean(avg_mean_tardiness_list)
        avg_max_tardiness = mean(avg_max_tardiness_list)
        avg_total_tardiness = mean(avg_total_tardiness_list)
        avg_number_tardy_jobs = mean(avg_number_tardy_jobs_list)
        performance = [avg_makespan, avg_number_tardy_jobs, avg_total_tardiness]
        lambda_performance = [0.33, 0.33, 0.33]
        reference_point = reference_point_input
        roh = 0.0001
        # calculate the fitness value according to the achievement scalarizing function
        if ref_point==None:
            fitness = performance[performance_measure]
        else:
            fitness = max(lambda_performance[i]*(performance[i]-reference_point[i]) for i in range(0,len(reference_point))) + roh * sum(lambda_performance[i]*(performance[i]-reference_point[i]) for i in range(0,len(reference_point)))
        return fitness,


    toolbox.register("evaluate", evalSymbReg)
    toolbox.register("select", tools.selTournament, tournsize=5)
    toolbox.register("mate", gp.cxOnePoint)
    toolbox.register("expr_mut", gp.genFull, min_=0, max_=2)
    toolbox.register("mutate", gp.mutUniform, expr=toolbox.expr_mut, pset=pset)

    toolbox.decorate("mate", gp.staticLimit(key=operator.attrgetter("height"), max_value=max_depth_crossover))
    toolbox.decorate("mutate", gp.staticLimit(key=operator.attrgetter("height"), max_value=max_depth_mutation))

    #random.seed(318)

    pop = toolbox.population(n=population_size)
    hof = tools.HallOfFame(1)

    stats_fit = tools.Statistics(lambda ind: ind.fitness.values)
    stats_size = tools.Statistics(len)
    mstats = tools.MultiStatistics(fitness=stats_fit, size=stats_size)
    mstats.register("avg", np.mean)
    mstats.register("std", np.std)
    mstats.register("min", np.min)
    mstats.register("max", np.max)

    pop, log = algorithms.eaSimple(pop, toolbox, crossover_probability, mutation_probability, nb_generations, stats=mstats,
                                   halloffame=hof, verbose=True)

    # extract statistics:
    avgFitnessValues  = log.chapters['fitness'].select("avg")
    minFitnessValues = log.chapters['fitness'].select("min")
    maxFitnessValues = log.chapters['fitness'].select("max")
    stdFitnessValues = log.chapters['fitness'].select("std")
    nb_generation = log.select("gen")
    nevals = log.select('nevals')

    # plot statistics:
    # sns.set_style("whitegrid")
    fig, ax = plt.subplots()
    #ax.boxplot(minFitnessValues)
    mins = np.array(minFitnessValues)
    maxes = np.array(maxFitnessValues)
    means = np.array(avgFitnessValues)
    std = np.array(stdFitnessValues)
    gen = np.array(nb_generation)

    # create stacked errorbars:

    #ax.scatter(nb_generation, avgFitnessValues)
    #ax.errorbar(gen, means, std, fmt='ok', lw=3)
    ax.errorbar(gen, means, [means - mins, maxes - means],
                 fmt='.k', ecolor='gray', lw=1)
    ax.set_xlabel('Generation')
    ax.set_ylabel('Average Fitness')
    st.header('Evolution process')
    st.pyplot(fig)

    #print("best solution:")
    #print(hof[0])
    best = hof.items[0]
    best_fitness = best.fitness.values[0]
    st.header('Dispatching rule of the best solution')
    st.write(str(best))

    # create visualisation of tree of the final solution
    nodes, edges, labels = gp.graph(hof.items[0])
    fig, ax = plt.subplots()
    g = nx.Graph()
    g.add_nodes_from(nodes)
    g.add_edges_from(edges)
    pos = nx.nx_agraph.graphviz_layout(g, prog='dot')
    nx.draw_networkx_nodes(g, pos)
    nx.draw_networkx_edges(g, pos)
    nx.draw_networkx_labels(g, pos, labels, font_size=6, font_color="whitesmoke")
    path = 'D:/PycharmProjects/05_DSS_hyper_heuristic/hyperheuristics/html_files'
    plt.savefig(f'{path}/nx_graph.png')
    image = Image.open(f'{path}/nx_graph.png')
    st.header('Primitive tree of the best solution')
    st.image(image,  use_column_width=True)

    wb = op.load_workbook(project_name + '.xlsx')
    ws_ideal_nadir =  wb["ideal_nadir"]
    ws_ref_point = wb["ref_point"]
    max_row = ws_ref_point.max_row
    if ref_point==None:
        st.header(performance_name[performance_measure])
        st.write(opt_type + " solution")
        st.write(best_fitness)
        if opt_type == 'best':
            ws_ideal_nadir['B'+str(performance_measure+2)] = best_fitness
        else:
            ws_ideal_nadir['C' + str(performance_measure + 2)] = best_fitness
    else:
        st.header('Best solution')
        st.write(best_fitness)
        ws_ref_point['A' + str(max_row + 1)] = str(reference_point_input)
        #ws_ref_point['A'+str(max_row+1)] = "["+str(reference_point_input[0]) +", " + str(reference_point_input[1]) +", "+ str(reference_point_input[2]) +"]"
        ws_ref_point['B' + str(max_row + 1)] = best_fitness
        ws_ref_point['C' + str(max_row + 1)] = str(best)

    print('best fitness slow \t\t', best_fitness)

    wb.save(project_name+'.xlsx')
    wb.close()
    #st.write(pop)
    #st.write(log)
    #st.stop()

def geneticprogamming_speedup_benchmark(performance_measure, ref_point, opt_type, project_name, reference_point_input, nb_generations, population_size, crossover_probability, mutation_probability, max_depth_crossover, max_depth_mutation, nb_simulations, lambda_performance):
    # define permormance measure names
    performance_name = ["makespan", "number tardy jobs", "total tardiness"]

    # define the functions and terminals
    pset = gp.PrimitiveSet("MAIN", 6)
    pset.addPrimitive(operator.add, 2)
    pset.addPrimitive(operator.sub, 2)
    pset.addPrimitive(operator.mul, 2)
    pset.addPrimitive(min, 2)
    pset.addPrimitive(max, 2)
    pset.addPrimitive(div, 2)
    pset.addPrimitive(ifte, 3)
    # there is currently no constant value as function defined
    # maybe change here to random between 0 and 10 as mentioned in overleaf

    # Definitions of the abbreviations:
    # PT: Processing Time of an operation
    # RPT: Remaining processing time of a job
    # RNO: Remaining number of uncompleted operations of a job
    # DD: Due Date of a job
    # SPTQ: Sum of processing time in queue
    # APTQ: Average processing time in queue
    # MAXPTQ: Maximum processing time in queue
    # MINPTQ: Minimum processing time in queue
    # MAXDD: Maximum Due date of remaining jobs
    # NRJ: Number of remaining jobs
    # SPT: Sum of processing time of all waiting jobs in the problem
    # TRNO: Total remaining number of uncompleted operations in the problem
    # CT: Current time

    # rename the terminals
    pset.renameArguments(ARG0='PT')
    pset.renameArguments(ARG1='RPT')
    pset.renameArguments(ARG2='RNO')
    pset.renameArguments(ARG3='DD')
    pset.renameArguments(ARG4='SPT')
    pset.renameArguments(ARG5='CT')

    # define if the problem is to be minimized or maximized (fitness minimization or maximization)
    if opt_type=="best":
        creator.create("FitnessMin", base.Fitness, weights=(-1.0,))
        creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMin)
    else:
        creator.create("FitnessMax", base.Fitness, weights=(1.0,))
        creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMax)

    # set some GP parameters
    toolbox = base.Toolbox()
    toolbox.register("expr", gp.genHalfAndHalf, pset=pset, min_=2, max_=6)
    toolbox.register("individual", tools.initIterate, creator.Individual, toolbox.expr)
    toolbox.register("population", tools.initRepeat, list, toolbox.individual)
    toolbox.register("compile", gp.compile, pset=pset)

    def evalSymbReg(individual):
        func = toolbox.compile(expr=individual)

        #def source(env, number, machine):
            #"""Source generates jobs randomly"""
            #for i in range(number):
            #    processing_time = np.random.uniform(avg_processing_time - delta_processing_time,
            #                                        avg_processing_time + delta_processing_time)
            #    if i == 0:
            #        release_time = np.random.uniform(avg_job_interarrival_time - delta_release_date,
            #                                         avg_job_interarrival_time + delta_release_date)
            #    else:
            #        release_time = release_time_list[i - 1] + np.random.uniform(
            #            avg_job_interarrival_time - delta_release_date,
            #            avg_job_interarrival_time + delta_release_date)

            #    due_date = release_time + k * processing_time
            #    due_date_list.append(due_date)
            #    processing_time_list.append(processing_time)
            #    release_time_list.append(release_time)
            #    c = job(env, f'Job {i + 1}', machine, processing_time=processing_time,
            #            total_processing_time=total_processing_time)
            #    env.process(c)
            #    t = 0
            #    yield env.timeout(t)

        def job(env, name, machine, PT, total_processing_time, RPT, DD,
                release_date, RNO):
            """Job arrives, is served and leaves."""
            CT = env.now
            SPT=1
            with machine.request(priority=func(PT, RPT, RNO, DD, SPT, CT)) as req:
                yield req
                # wait = env.now - arrive # waiting time of job
                #job_start = env.now

                # We got to the counter
                # print('%7.4f %s: Waited %6.3f' % (env.now, name, wait))

                yield env.timeout(PT)
                # print('%7.4f %s: Finished' % (env.now, name))

                # Flow Time
                # flow_time.append(env.now-arrive)

                job_prec = name
                machine_prec = machine_list.index(machine)
                schedule.append([job_prec, machine_prec])

                if name==1:
                    print('job nr.: ' + str(name))
                    print('current time: ' + str(CT))
                    print('remaining processing time before operation: ' + str(RPT))
                    print('number of remaining operations before operation: ' + str(RNO))
                    print('Due date of job: '+str(DD))

                RPT -= PT
                RNO -= 1

                if name==1:
                    print('processing time operation: ' + str(PT))
                    print('remaining processing after operation: ' + str(RPT))
                    print('number of remaining operations after operation: ' + str(RNO))

                #results.append({'Job': f'J{job_prec}',
                #                'Machine': f'M{machine_prec}',
                #                'Start': job_start,
                #                'Duration': processing_time,
                #                'Finish': env.now})
                job_finished = 'Yes'
                for j in range(n_jobs):
                    for m in range(number_machines):
                        if TASKS[j, m]['prec'] == (job_prec, machine_prec):
                            machine = machine_list[m]
                            PT = float(TASKS[j, m]['dur'])
                            job_finished = 'No'
                            env.process(
                                job(env, j, machine, PT, total_processing_time, RPT,
                                    DD, release_date, RNO))
                if job_finished == 'Yes':
                    # Completion time
                    completion_time.append(env.now)
                    # Tardiness of job
                    tardiness.append(max(env.now - DD, 0))
                    # Tardy jobs
                    if max(env.now - DD, 0) > 0:
                        tardy_jobs.append(1)
                    else:
                        tardy_jobs.append(0)

        avg_makespan_list = []
        avg_mean_tardiness_list = []
        avg_max_tardiness_list = []
        avg_total_tardiness_list = []
        avg_number_tardy_jobs_list = []

        for simulations in range(nb_simulations):
            # load the jobs with its corresponding data
            TASKS, total_processing_time, due_date_list_jobs, release_date_list, n_jobs, number_machines, SPT = load_test_problem_random(
                name='random')

            schedule = []
            tardiness = []
            tardy_jobs = []
            completion_time = []

            env = simpy.Environment()

            # Start processes and run
            machine_0 = simpy.PriorityResource(env, capacity=1)
            machine_1 = simpy.PriorityResource(env, capacity=1)
            machine_2 = simpy.PriorityResource(env, capacity=1)
            machine_3 = simpy.PriorityResource(env, capacity=1)
            machine_4 = simpy.PriorityResource(env, capacity=1)
            machine_5 = simpy.PriorityResource(env, capacity=1)
            machine_6 = simpy.PriorityResource(env, capacity=1)
            machine_7 = simpy.PriorityResource(env, capacity=1)
            machine_8 = simpy.PriorityResource(env, capacity=1)
            machine_9 = simpy.PriorityResource(env, capacity=1)
            machine_10 = simpy.PriorityResource(env, capacity=1)
            machine_11 = simpy.PriorityResource(env, capacity=1)
            machine_12 = simpy.PriorityResource(env, capacity=1)
            machine_13 = simpy.PriorityResource(env, capacity=1)
            machine_14 = simpy.PriorityResource(env, capacity=1)
            # machine = simpy.Resource(env, capacity=1)

            machine_list = [machine_0, machine_1, machine_2, machine_3, machine_4, machine_5, machine_6, machine_7,
                            machine_8, machine_9, machine_10, machine_11, machine_12, machine_13, machine_14]

            for j in range(n_jobs):
                for m in range(number_machines):
                    if TASKS[j, m]['prec'] == None:
                        current_machine = machine_list[m]
                        PT = float(TASKS[j, m]['dur'])
                        job_new = j
                        total_processing_time_current_job = total_processing_time[j]
                        RPT = total_processing_time_current_job
                        release_date_current_job = release_date_list[j]
                        DD = due_date_list_jobs[j]
                        RNO = number_machines
                        env.process(job(env, job_new, current_machine, PT=PT,
                                        total_processing_time=total_processing_time_current_job,
                                        RPT=RPT,
                                        DD=DD, release_date=release_date_current_job,
                                        RNO=RNO))

            env.run()

            # for i in range(number_simulations):
            #    env.process(source(env, n_jobs, machine_1))
            #    env.run()

            # Post processing
            # calculate and performance measures of the current simulation run
            total_tardiness = sum(tardiness)
            mean_tardiness = mean(tardiness)
            max_tardiness = max(tardiness)
            number_tardy_jobs = sum(tardy_jobs)
            makespan = max(completion_time)
            # mean_flow_time = mean(flow_time)
            # max_flow_time = max(flow_time)
            # print('Release Time')
            # print(release_time_list)
            # print('processing time')
            # print(processing_time_list)
            # print('Due Dates')
            # print(due_date_list)
            # print(f'Total Tardiness: {total_tardiness}')
            # print(f'Mean Tardiness: {mean_tardiness}')
            # print(f'Max Tardiness: {max_tardiness}')
            # print(f'Number of tardy Jobs: {number_tardy_jobs}')
            # print(completion_time)
            # print(f'Makespan: {makespan}')
            # print(f'Mean flow time: {mean_flow_time}')
            # print(f'Max flow time: {max_flow_time}')
            # print(results)

            # add performance measures of current simulation run to the list for all runs
            avg_makespan_list.append(makespan)
            avg_mean_tardiness_list.append(mean_tardiness)
            avg_max_tardiness_list.append(max_tardiness)
            avg_total_tardiness_list.append(total_tardiness)
            avg_number_tardy_jobs_list.append(number_tardy_jobs)

            # visualize(results)

        # calculate and print the performance measures after all simulation runs
        avg_makespan = mean(avg_makespan_list)
        avg_mean_tardiness = mean(avg_mean_tardiness_list)
        avg_max_tardiness = mean(avg_max_tardiness_list)
        avg_total_tardiness = mean(avg_total_tardiness_list)
        avg_number_tardy_jobs = mean(avg_number_tardy_jobs_list)
        performance = [avg_makespan, avg_number_tardy_jobs, avg_total_tardiness]
        reference_point = reference_point_input
        roh = 0.0001
        # calculate the fitness value according to the achievement scalarizing function
        if ref_point==None:
            fitness = performance[performance_measure]
        else:
            fitness = max(lambda_performance[i]*(performance[i]-reference_point[i]) for i in range(0,len(reference_point))) + roh * sum(lambda_performance[i]*(performance[i]-reference_point[i]) for i in range(0,len(reference_point)))
        return fitness,

    # initialize GP and set some parameter
    toolbox.register("evaluate", evalSymbReg)
    toolbox.register("select", tools.selTournament, tournsize=5)
    toolbox.register("mate", gp.cxOnePoint)
    toolbox.register("expr_mut", gp.genFull, min_=0, max_=2)
    toolbox.register("mutate", gp.mutUniform, expr=toolbox.expr_mut, pset=pset)
    toolbox.decorate("mate", gp.staticLimit(key=operator.attrgetter("height"), max_value=max_depth_crossover))
    toolbox.decorate("mutate", gp.staticLimit(key=operator.attrgetter("height"), max_value=max_depth_mutation))
    #random.seed(318)

    # define population and hall of fame (size of the best kept solutions found during GP run)
    pop = toolbox.population(n=population_size)
    hof = tools.HallOfFame(1)

    # define statistics for the GP run to be measured
    stats_fit = tools.Statistics(lambda ind: ind.fitness.values)
    stats_size = tools.Statistics(len)
    mstats = tools.MultiStatistics(fitness=stats_fit, size=stats_size)
    mstats.register("avg", np.mean)
    mstats.register("std", np.std)
    mstats.register("min", np.min)
    mstats.register("max", np.max)

    # get population and logbook from the whole GP run
    pop, log = algorithms.eaSimple(pop, toolbox, crossover_probability, mutation_probability, nb_generations, stats=mstats,
                                   halloffame=hof, verbose=True)

    # extract statistics:
    avgFitnessValues  = log.chapters['fitness'].select("avg")
    minFitnessValues = log.chapters['fitness'].select("min")
    maxFitnessValues = log.chapters['fitness'].select("max")
    stdFitnessValues = log.chapters['fitness'].select("std")
    nb_generation = log.select("gen")
    nevals = log.select('nevals')

    # transform statistics into numpy arrays
    minFitnessValues = np.array(minFitnessValues)
    maxFitnessValues = np.array(maxFitnessValues)
    avgFitnessValues = np.array(avgFitnessValues)
    stdFitnessValues = np.array(stdFitnessValues)
    nb_generation = np.array(nb_generation)

    # load best solution of the GP run
    best_solution = hof.items[0]
    best_fitness = best_solution.fitness.values[0]
    return best_solution, best_fitness, nb_generation, avgFitnessValues, minFitnessValues, maxFitnessValues, stdFitnessValues

def geneticprogamming_simulation(performance_measure, ref_point, opt_type, project_name, reference_point_input, nb_generations, population_size, crossover_probability, mutation_probability, max_depth_crossover, max_depth_mutation, nb_simulations, lambda_performance, constraint_index, constraint_value):
    # define permormance measure names
    performance_name = ["makespan", "number tardy jobs", "total tardiness"]

    # define the functions and terminals
    pset = gp.PrimitiveSet("MAIN", 13)
    pset.addPrimitive(operator.add, 2)
    pset.addPrimitive(operator.sub, 2)
    pset.addPrimitive(operator.mul, 2)
    pset.addPrimitive(min, 2)
    pset.addPrimitive(max, 2)
    pset.addPrimitive(div, 2)
    pset.addPrimitive(ifte, 3)
    # there is currently no constant value as function defined
    #randomname = str(random.randint(1, 100000000000000)) + str(random.randint(1, 100000000000000)) + str(random.randint(1, 100000000000000))
    # Problem bei Konstante da jedes mal neuer Name generiert werden muss
    #try:
    #    pset.addEphemeralConstant(randomname, lambda: random.randint(0, 10))
    #except:
    #    randomname = str(random.randint(1, 100000000000000)) + str(random.randint(1, 100000000000000)) + str(
    #        random.randint(1, 100000000000000))
    #    pset.addEphemeralConstant(randomname, lambda: random.randint(0, 10))

    # maybe change here to random between 0 and 10 as mentioned in overleaf

    # Definitions of the abbreviations:
    # PT: Processing Time of an operation
    # RPT: Remaining processing time of a job
    # RNO: Remaining number of uncompleted operations of a job
    # DD: Due Date of a job
    # SPTQ: Sum of processing time in queue
    # APTQ: Average processing time in queue
    # MAXPTQ: Maximum processing time in queue
    # MINPTQ: Minimum processing time in queue
    # MAXDD: Maximum Due date of remaining jobs
    # NRJ: Number of remaining jobs
    # SPT: Sum of processing time of all waiting jobs in the problem
    # TRNO: Total remaining number of uncompleted operations in the problem
    # CT: Current time

    # rename the terminals
    pset.renameArguments(ARG0='PT')
    pset.renameArguments(ARG1='RPT')
    pset.renameArguments(ARG2='RNO')
    pset.renameArguments(ARG3='DD')
    pset.renameArguments(ARG4='SPTQ')
    pset.renameArguments(ARG5='APTQ')
    pset.renameArguments(ARG6='MAXPTQ')
    pset.renameArguments(ARG7='MINPTQ')
    pset.renameArguments(ARG8='MAXDDQ')
    pset.renameArguments(ARG9='NJQ')
    pset.renameArguments(ARG10='SPT')
    pset.renameArguments(ARG11='TRNO')
    pset.renameArguments(ARG12='CT')

    # define if the problem is to be minimized or maximized (fitness minimization or maximization)
    if opt_type=="best":
        creator.create("FitnessMin", base.Fitness, weights=(-1.0,))
        creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMin)
    else:
        creator.create("FitnessMax", base.Fitness, weights=(1.0,))
        creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMax)

    # set some GP parameters
    toolbox = base.Toolbox()
    toolbox.register("expr", gp.genHalfAndHalf, pset=pset, min_=2, max_=6)
    toolbox.register("individual", tools.initIterate, creator.Individual, toolbox.expr)
    toolbox.register("population", tools.initRepeat, list, toolbox.individual)
    toolbox.register("compile", gp.compile, pset=pset)

    def evalSymbReg(individual):
        func = toolbox.compile(expr=individual)

        #print(individual)

        results = []
        makespan = []
        tardiness = []
        waiting_time = []
        random_seed = [0, 10, 15, 36, 27, 50, 82, 99, 103, 230] # train set
        #random_seed = [5, 20, 30, 42, 51, 65, 87, 101, 110, 242] # test set
        '''
        pool = mp.Pool(mp.cpu_count())
        results, makespan, tardiness, earliness = pool.starmap(simulation, [(15, 20, func, random_seed[simulations]) for simulations in range(nb_simulations)])
        pool.close()
        '''

        for simulations in range(nb_simulations):
            # perform simulation
            results_current_simulation_run, makespan_current_simulation_run, tardiness_current_simulation_run, waiting_time_current_simulation_run = \
                simulation(number_machines=15, number_jobs=20, func=func, random_seed=random_seed[simulations])
            #results.append(results_current_simulation_run)
            makespan.append(makespan_current_simulation_run)
            tardiness.append(tardiness_current_simulation_run)
            waiting_time.append(waiting_time_current_simulation_run)

        # calculate and print the performance measures after all simulation runs
        avg_makespan = np.mean(makespan)
        avg_total_tardiness = np.mean(tardiness)
        avg_waiting_time = np.mean(waiting_time)
        performance = [avg_makespan, avg_total_tardiness, avg_waiting_time]
        reference_point = reference_point_input
        roh = 0.0001
        # calculate the fitness value according to the achievement scalarizing function
        if ref_point==None:
            fitness = performance[performance_measure]
        else:
            fitness = max(lambda_performance[i]*(performance[i]-reference_point[i]) for i in range(0,len(reference_point))) + roh * sum(lambda_performance[i]*(performance[i]-reference_point[i]) for i in range(0,len(reference_point)))
        # check if the value force against the constraint and punish it in case of
        if constraint_index != None:
            if performance[constraint_index] > constraint_value:
                fitness += 100
        return fitness,

    # initialize GP and set some parameter
    toolbox.register("evaluate", evalSymbReg)
    toolbox.register("select", tools.selTournament, tournsize=5)
    toolbox.register("mate", gp.cxOnePoint)
    toolbox.register("expr_mut", gp.genFull, min_=0, max_=2)
    toolbox.register("mutate", gp.mutUniform, expr=toolbox.expr_mut, pset=pset)
    toolbox.decorate("mate", gp.staticLimit(key=operator.attrgetter("height"), max_value=max_depth_crossover))
    toolbox.decorate("mutate", gp.staticLimit(key=operator.attrgetter("height"), max_value=max_depth_mutation))
    #random.seed(318)

    # define population and hall of fame (size of the best kept solutions found during GP run)
    pop = toolbox.population(n=population_size)
    hof = tools.HallOfFame(1)

    # define statistics for the GP run to be measured
    stats_fit = tools.Statistics(lambda ind: ind.fitness.values)
    stats_size = tools.Statistics(len)
    mstats = tools.MultiStatistics(fitness=stats_fit, size=stats_size)
    mstats.register("avg", np.mean)
    mstats.register("std", np.std)
    mstats.register("min", np.min)
    mstats.register("max", np.max)

    # get population and logbook from the whole GP run
    pop, log = algorithms.eaSimple(pop, toolbox, crossover_probability, mutation_probability, nb_generations, stats=mstats,
                                   halloffame=hof, verbose=True)

    # extract statistics:
    avgFitnessValues  = log.chapters['fitness'].select("avg")
    minFitnessValues = log.chapters['fitness'].select("min")
    maxFitnessValues = log.chapters['fitness'].select("max")
    stdFitnessValues = log.chapters['fitness'].select("std")
    nb_generation = log.select("gen")
    nevals = log.select('nevals')

    # transform statistics into numpy arrays
    minFitnessValues = np.array(minFitnessValues)
    maxFitnessValues = np.array(maxFitnessValues)
    avgFitnessValues = np.array(avgFitnessValues)
    stdFitnessValues = np.array(stdFitnessValues)
    nb_generation = np.array(nb_generation)

    # load best solution of the GP run
    best_solution = hof.items[0]
    best_fitness = best_solution.fitness.values[0]
    return best_solution, best_fitness, nb_generation, avgFitnessValues, minFitnessValues, maxFitnessValues, stdFitnessValues

def geneticprogamming_verification(performance_measure, ref_point, opt_type, project_name, reference_point_input, nb_generations, population_size, crossover_probability, mutation_probability, max_depth_crossover, max_depth_mutation, nb_simulations):
    performance_name = ["makespan", "number tardy jobs", "total tardiness"]
    def div(left, right):
        try:
            return left / right
        except ZeroDivisionError:
            return 1

    pset = gp.PrimitiveSet("MAIN", 4)
    pset.addPrimitive(operator.add, 2)
    pset.addPrimitive(operator.sub, 2)
    pset.addPrimitive(operator.mul, 2)
    pset.addPrimitive(min, 2)
    pset.addPrimitive(max, 2)
    pset.addPrimitive(div, 2)
    #pset.addPrimitive(operator.neg, 1)
    #pset.addPrimitive(math.cos, 1)
    #pset.addPrimitive(math.sin, 1)
    #randomname = str(rd.randint(1, 100000000000000)) + str(rd.randint(1, 100000000000000)) + str(rd.randint(1, 100000000000000))
    # Problem bei Konstante da jedes mal neuer Name generiert werden muss
    #try:
    #    pset.addEphemeralConstant(randomname, lambda: rd.uniform(-1, 1))
    #except:
    #    randomname = str(rd.randint(1, 100000000000000)) + str(rd.randint(1, 100000000000000)) + str(
    #        rd.randint(1, 100000000000000))
    #    pset.addEphemeralConstant(randomname, lambda: rd.uniform(-1, 1))
    #maybe change here to random between 0 and 10 as mentioned in overleaf
    pset.renameArguments(ARG0='PT')
    pset.renameArguments(ARG1='SPT')
    pset.renameArguments(ARG2='RPT')
    pset.renameArguments(ARG3='DD')

    if opt_type=="best":
        creator.create("FitnessMin", base.Fitness, weights=(-1.0,))
        creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMin)
    else:
        creator.create("FitnessMax", base.Fitness, weights=(1.0,))
        creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMax)



    toolbox = base.Toolbox()
    toolbox.register("expr", gp.genHalfAndHalf, pset=pset, min_=2, max_=6)
    toolbox.register("individual", tools.initIterate, creator.Individual, toolbox.expr)
    toolbox.register("population", tools.initRepeat, list, toolbox.individual)
    toolbox.register("compile", gp.compile, pset=pset)


    # random job generator
    '''np.random.seed(10)
    number_machines, n_jobs, duedate_tightness = 10, 100, 1.5
    TASKS = {}
    release_date_list, due_date_list_jobs, total_processing_time = [], [], []
    for i in range(n_jobs):
        prec = None
        release_time = np.random.uniform(0, 40)
        sum_proc_time = 0
        allowed_values = list(range(0, 10))
        for m in range(number_machines):
            dur = np.random.uniform(number_machines / 2, number_machines * 2)
            sum_proc_time += dur
            machine = np.random.choice(allowed_values)
            task = (i, machine)
            TASKS[task] = {'dur': float(dur), 'prec': prec}
            prec = task
            allowed_values.remove(machine)
        due_date = release_time + duedate_tightness * sum_proc_time
        release_date_list.append(release_time)
        due_date_list_jobs.append(due_date)
        total_processing_time.append(sum_proc_time)'''

    # TA15 benchmark problem
    number_machines, n_jobs = 15, 20
    data = """
 7 84  0 58 12 71  4 26  1 98  9 36  2 12 11 30 10 87 14 95  5 45  6 28 13 73  3 73  8 45 
 4 29  8 22  7 47  3 75  9 94 13 15 12  4  0 82 11 14 10 35  1 79  6 34  5 57 14 23  2 56 
 1 73  4 36  7 48 13 26  3 49  8 60 10 15  5 66 12 90 14 39  9  8  6 74  2 63  0 94 11 91 
 5  1 11 35  9 23 12 93  7 75  1 50  6 40 13 60  8 41  2  7  0 57 14 72  3 40  4 75 10  7 
 4 13 11 15 12 17  1 14  0 67  9 94  6 18 13 52  2 53 14 16  5 33 10 61  3 47  8 65  7 39 
 2 54  6 80  3 87  8 36 14 54  0 72  4 17 10 44 11 37  1 88  7 77 13 84 12 17  5 82  9 90 
 4  4 14 62  5 33 10 62  8 86  7 30  6 39  1 67  0 42 12 31  9 83 13 39 11 67  3 67  2 31 
 7 29 10 29 11 69 14 26  3 55  2 46  4 53  5 65  1 97 12 24  9 69  6 22 13 17  0 39  8 13 
14 12 11 73  0 36 13 70  3 12  2 80  1 99  8 70  5 51  7 14  4 71 12 28  6 35 10 58  9 35 
 0 61  5 49 12 74  1 90 13 60 10 88  9  3  4 60  2 59  8 94 14 91 11 34  7 26  6  4  3 26 
 4 89  3 90  8 95 12 32  9 18 11 73  2  9 14 19  5 97  7 58 13 36  6 62 10 13  1 16  0  1 
 9 71  6 47  1 95  0  7 14 63  7 49 13 24 12 46  2 72 11 73  5 19  8 96 10 41  3 15  4 81 
 4 45  3  9  0 97 14 62 13 77  9 78  7 70  2 19 11 86  8 15 10 23  1 46  6 32 12  6  5 70 
12 74 10 46  3 98  6  1  4 53  5 59  0 86  7 98  2 76  8 12 13 91 11 98 14 98  9 11  1 27 
14 73  7 70  5 14  8 32 11 19  0 57  2 17 13 96 12 56  4 73  6 32  1  7 10 79  9 10  3 91 
 6 39 14 87 12 11  2 81  7  7  5 79  8 24 13  9 11 58  9 42  0 67  3 27  4 20  1 19 10 67 
 9 76  5 89 14 64 10 14 12 11  1 14  4 99 13 85  0 81 11  3  3 46  2 47  7 40  6 81  8 27 
 9 55 12 71  4  5 14 83 11 16  8  4  0 20  7 15  5 60  3  8  1 93 10 33  6 63 13 71  2 29 
12 92  2 25  3  8 14 86  5 22  1 79  6 23 11 96 13 24  9 94  7 97 10 17  8 48  0 67  4 47 
 3  5 12 77 10 74  5 59 14 13  0 57  9 62  8 37 13 54  6 69 11 80  1 35  7 88  2 47  4 98 
     """
    TASKS = {}
    for job, line in enumerate(data.splitlines()[1:]):
        nums = line.split()
        prec = None
        for m, dur in zip(nums[::2], nums[1::2]):
            task = (job, int(m))
            TASKS[task] = {'dur': int(dur), 'prec': prec}
            prec = task

    release_date_list = []
    due_date_list_jobs = []
    total_processing_time = []
    #print(TASKS)

    for j in range(n_jobs):
        total_processing_time_current_job = 0
        for m in range(number_machines):
            total_processing_time_current_job += TASKS[j, m]['dur']
        total_processing_time.append(total_processing_time_current_job)
        due_date_list_jobs.append(1.3 * total_processing_time_current_job)

    def evalSymbReg(individual):
        func = toolbox.compile(expr=individual)

        def visualize(results):
            schedule = pd.DataFrame(results)
            JOBS = sorted(list(schedule['Job'].unique()))
            MACHINES = sorted(list(schedule['Machine'].unique()))
            makespan = schedule['Finish'].max()

            bar_style = {'alpha': 1.0, 'lw': 25, 'solid_capstyle': 'butt'}
            text_style = {'color': 'white', 'weight': 'bold', 'ha': 'center', 'va': 'center'}
            colors = mpl.cm.Dark2.colors

            schedule.sort_values(by=['Job', 'Start'])
            schedule.set_index(['Job', 'Machine'], inplace=True)

            fig, ax = plt.subplots(2, 1, figsize=(12, 5 + (len(JOBS) + len(MACHINES)) / 4))

            for jdx, j in enumerate(JOBS, 1):
                for mdx, m in enumerate(MACHINES, 1):
                    if (j, m) in schedule.index:
                        xs = schedule.loc[(j, m), 'Start']
                        xf = schedule.loc[(j, m), 'Finish']
                        ax[0].plot([xs, xf], [jdx] * 2, c=colors[mdx % 7], **bar_style)
                        ax[0].text((xs + xf) / 2, jdx, m, **text_style)
                        ax[1].plot([xs, xf], [mdx] * 2, c=colors[jdx % 7], **bar_style)
                        ax[1].text((xs + xf) / 2, mdx, j, **text_style)

            ax[0].set_title('Job Schedule')
            ax[0].set_ylabel('Job')
            ax[1].set_title('Machine Schedule')
            ax[1].set_ylabel('Machine')

            for idx, s in enumerate([JOBS, MACHINES]):
                ax[idx].set_ylim(0.5, len(s) + 0.5)
                ax[idx].set_yticks(range(1, 1 + len(s)))
                ax[idx].set_yticklabels(s)
                ax[idx].text(makespan, ax[idx].get_ylim()[0] - 0.2, "{0:0.1f}".format(makespan), ha='center', va='top')
                ax[idx].plot([makespan] * 2, ax[idx].get_ylim(), 'r--')
                ax[idx].set_xlabel('Time')
                ax[idx].grid(True)

            fig.tight_layout()
            plt.show()

        def source(env, number, machine):
            """Source generates jobs randomly"""
            for i in range(number):
                processing_time = np.random.uniform(avg_processing_time - delta_processing_time,
                                                    avg_processing_time + delta_processing_time)
                if i == 0:
                    release_time = np.random.uniform(avg_job_interarrival_time - delta_release_date,
                                                     avg_job_interarrival_time + delta_release_date)
                else:
                    release_time = release_time_list[i - 1] + np.random.uniform(
                        avg_job_interarrival_time - delta_release_date,
                        avg_job_interarrival_time + delta_release_date)

                due_date = release_time + k * processing_time
                due_date_list.append(due_date)
                processing_time_list.append(processing_time)
                release_time_list.append(release_time)
                c = job(env, f'Job {i + 1}', machine, processing_time=processing_time,
                        total_processing_time=total_processing_time)
                env.process(c)
                t = 0
                yield env.timeout(t)

        def job(env, name, machine, processing_time, total_processing_time, remaining_processing_time, due_date,
                release_date):
            """Job arrives, is served and leaves."""
            # arrive = release_time
            # print('%7.4f %s: Arrived' % (arrive, name))

            with machine.request(priority=func(processing_time, total_processing_time, remaining_processing_time, due_date)) as req:
                yield req
                # wait = env.now - arrive # waiting time of job
                job_start = env.now

                # We got to the counter
                # print('%7.4f %s: Waited %6.3f' % (env.now, name, wait))

                yield env.timeout(processing_time)
                # print('%7.4f %s: Finished' % (env.now, name))

                # Flow Time
                # flow_time.append(env.now-arrive)
                remaining_processing_time -= processing_time
                job_prec = name
                machine_prec = machine_list.index(machine)
                schedule.append([job_prec, machine_prec])
                #results.append({'Job': f'J{job_prec}',
                #                'Machine': f'M{machine_prec}',
                #                'Start': job_start,
                #                'Duration': processing_time,
                #                'Finish': env.now})
                job_finished = 'Yes'
                for j in range(n_jobs):
                    for m in range(number_machines):
                        if TASKS[j, m]['prec'] == (job_prec, machine_prec):
                            machine = machine_list[m]
                            processing_time = TASKS[j, m]['dur']
                            job_finished = 'No'
                            env.process(
                                job(env, j, machine, processing_time, total_processing_time, remaining_processing_time,
                                    due_date, release_date))
                if job_finished == 'Yes':
                    # Completion time
                    completion_time.append(env.now)
                    # Tardiness of job
                    tardiness.append(max(env.now - due_date, 0))
                    # Tardy jobs
                    if max(env.now - due_date, 0) > 0:
                        tardy_jobs.append(1)
                    else:
                        tardy_jobs.append(0)

        avg_makespan_list = []
        avg_mean_tardiness_list = []
        avg_max_tardiness_list = []
        avg_total_tardiness_list = []
        avg_number_tardy_jobs_list = []

        for simulations in range(nb_simulations):




            # print(TASKS['J0', 'M0']['prec'])
            # print(pd.DataFrame(TASKS).T)

            number_simulations = 1
            R_processing_time = 0.4
            avg_processing_time = 10
            processing_time = []
            duedate_tightness = 2
            duedate_variability = 0.3
            machine_utilization = 0.7
            job_interarrival_tightness = 1
            schedule = []
            release_time = []
            tardiness = []
            tardy_jobs = []
            completion_time = []
            flow_time = []
            due_date_list = []
            release_time_list = []
            processing_time_list = []
            results = []
            remaining_processing_time_current_job = []

            n_jobs_original = n_jobs
            delta_processing_time = R_processing_time * avg_processing_time
            avg_job_interarrival_time = avg_processing_time / machine_utilization
            delta_release_date = job_interarrival_tightness * avg_job_interarrival_time
            release_time.append(np.random.uniform(avg_job_interarrival_time - delta_release_date,
                                                  avg_job_interarrival_time + delta_release_date))
            delta_duedate = duedate_tightness * duedate_variability
            k = np.random.uniform(duedate_tightness - delta_duedate, duedate_tightness + delta_duedate)

            env = simpy.Environment()

            # Start processes and run
            machine_0 = simpy.PriorityResource(env, capacity=1)
            machine_1 = simpy.PriorityResource(env, capacity=1)
            machine_2 = simpy.PriorityResource(env, capacity=1)
            machine_3 = simpy.PriorityResource(env, capacity=1)
            machine_4 = simpy.PriorityResource(env, capacity=1)
            machine_5 = simpy.PriorityResource(env, capacity=1)
            machine_6 = simpy.PriorityResource(env, capacity=1)
            machine_7 = simpy.PriorityResource(env, capacity=1)
            machine_8 = simpy.PriorityResource(env, capacity=1)
            machine_9 = simpy.PriorityResource(env, capacity=1)
            machine_10 = simpy.PriorityResource(env, capacity=1)
            machine_11 = simpy.PriorityResource(env, capacity=1)
            machine_12 = simpy.PriorityResource(env, capacity=1)
            machine_13 = simpy.PriorityResource(env, capacity=1)
            machine_14 = simpy.PriorityResource(env, capacity=1)
            # machine = simpy.Resource(env, capacity=1)

            machine_list = [machine_0, machine_1, machine_2, machine_3, machine_4, machine_5, machine_6, machine_7,
                            machine_8, machine_9, machine_10, machine_11, machine_12, machine_13, machine_14]

            for j in range(n_jobs):
                for m in range(number_machines):
                    if TASKS[j, m]['prec'] == None:
                        current_machine = machine_list[m]
                        processing_time_new = float(TASKS[j, m]['dur'])
                        job_new = j
                        total_processing_time_current_job = total_processing_time[j]
                        remaining_processing_time_current_job = total_processing_time_current_job
                        #release_date_current_job = release_date_list[j]
                        #due_date_current_job = due_date_list_jobs[j]
                        env.process(job(env, job_new, current_machine, processing_time=processing_time_new,
                                        total_processing_time=total_processing_time_current_job,
                                        remaining_processing_time=remaining_processing_time_current_job,
                                        due_date=1, release_date=1))

            env.run()

            # for i in range(number_simulations):
            #    env.process(source(env, n_jobs, machine_1))
            #    env.run()

            # Post processing
            # calculate and performance measures of the current simulation run
            total_tardiness = sum(tardiness)
            mean_tardiness = mean(tardiness)
            max_tardiness = max(tardiness)
            number_tardy_jobs = sum(tardy_jobs)
            makespan = max(completion_time)
            # mean_flow_time = mean(flow_time)
            # max_flow_time = max(flow_time)
            # print('Release Time')
            # print(release_time_list)
            # print('processing time')
            # print(processing_time_list)
            # print('Due Dates')
            # print(due_date_list)
            # print(f'Total Tardiness: {total_tardiness}')
            # print(f'Mean Tardiness: {mean_tardiness}')
            # print(f'Max Tardiness: {max_tardiness}')
            # print(f'Number of tardy Jobs: {number_tardy_jobs}')
            # print(completion_time)
            # print(f'Makespan: {makespan}')
            # print(f'Mean flow time: {mean_flow_time}')
            # print(f'Max flow time: {max_flow_time}')
            # print(results)

            # add performance measures of current simulation run to the list for all runs
            avg_makespan_list.append(makespan)
            avg_mean_tardiness_list.append(mean_tardiness)
            avg_max_tardiness_list.append(max_tardiness)
            avg_total_tardiness_list.append(total_tardiness)
            avg_number_tardy_jobs_list.append(number_tardy_jobs)

            # visualize(results)

        # calculate and print the performance measures after all simulation runs
        avg_makespan = mean(avg_makespan_list)
        avg_mean_tardiness = mean(avg_mean_tardiness_list)
        avg_max_tardiness = mean(avg_max_tardiness_list)
        avg_total_tardiness = mean(avg_total_tardiness_list)
        avg_number_tardy_jobs = mean(avg_number_tardy_jobs_list)
        performance = [avg_makespan, avg_number_tardy_jobs, avg_total_tardiness]
        lambda_performance = [0.33, 0.33, 0.33]
        reference_point = reference_point_input
        roh = 0.0001
        # calculate the fitness value according to the achievement scalarizing function
        if ref_point==None:
            fitness = performance[performance_measure]
        else:
            fitness = max(lambda_performance[i]*(performance[i]-reference_point[i]) for i in range(0,len(reference_point))) + roh * sum(lambda_performance[i]*(performance[i]-reference_point[i]) for i in range(0,len(reference_point)))
        return fitness,


    toolbox.register("evaluate", evalSymbReg)
    toolbox.register("select", tools.selTournament, tournsize=5)
    toolbox.register("mate", gp.cxOnePoint)
    toolbox.register("expr_mut", gp.genFull, min_=0, max_=2)
    toolbox.register("mutate", gp.mutUniform, expr=toolbox.expr_mut, pset=pset)

    toolbox.decorate("mate", gp.staticLimit(key=operator.attrgetter("height"), max_value=max_depth_crossover))
    toolbox.decorate("mutate", gp.staticLimit(key=operator.attrgetter("height"), max_value=max_depth_mutation))

    #random.seed(318)

    pop = toolbox.population(n=population_size)
    hof = tools.HallOfFame(1)

    stats_fit = tools.Statistics(lambda ind: ind.fitness.values)
    stats_size = tools.Statistics(len)
    mstats = tools.MultiStatistics(fitness=stats_fit, size=stats_size)
    mstats.register("avg", np.mean)
    mstats.register("std", np.std)
    mstats.register("min", np.min)
    mstats.register("max", np.max)

    pop, log = algorithms.eaSimple(pop, toolbox, crossover_probability, mutation_probability, nb_generations, stats=mstats,
                                   halloffame=hof, verbose=True)

    # extract statistics:
    avgFitnessValues  = log.chapters['fitness'].select("avg")
    minFitnessValues = log.chapters['fitness'].select("min")
    maxFitnessValues = log.chapters['fitness'].select("max")
    stdFitnessValues = log.chapters['fitness'].select("std")
    nb_generation = log.select("gen")
    nevals = log.select('nevals')

    # plot statistics:
    # sns.set_style("whitegrid")
    fig, ax = plt.subplots()
    #ax.boxplot(minFitnessValues)
    mins = np.array(minFitnessValues)
    maxes = np.array(maxFitnessValues)
    means = np.array(avgFitnessValues)
    std = np.array(stdFitnessValues)
    gen = np.array(nb_generation)

    # create stacked errorbars:

    #ax.scatter(nb_generation, avgFitnessValues)
    #ax.errorbar(gen, means, std, fmt='ok', lw=3)
    ax.errorbar(gen, means, [means - mins, maxes - means],
                 fmt='.k', ecolor='gray', lw=1)
    ax.set_xlabel('Generation')
    ax.set_ylabel('Average Fitness')
    st.header('Evolution process')
    st.pyplot(fig)

    #print("best solution:")
    #print(hof[0])
    best = hof.items[0]
    best_fitness = best.fitness.values[0]
    st.header('Dispatching rule of the best solution')
    st.write(str(best))

    # create visualisation of tree of the final solution
    nodes, edges, labels = gp.graph(hof.items[0])
    fig, ax = plt.subplots()
    g = nx.Graph()
    g.add_nodes_from(nodes)
    g.add_edges_from(edges)
    pos = nx.nx_agraph.graphviz_layout(g, prog='dot')
    nx.draw_networkx_nodes(g, pos)
    nx.draw_networkx_edges(g, pos)
    nx.draw_networkx_labels(g, pos, labels, font_size=6, font_color="whitesmoke")
    path = 'D:/PycharmProjects/05_DSS_hyper_heuristic/hyperheuristics/html_files'
    plt.savefig(f'{path}/nx_graph.png')
    image = Image.open(f'{path}/nx_graph.png')
    st.header('Primitive tree of the best solution')
    st.image(image,  use_column_width=True)

    wb = op.load_workbook(project_name + '.xlsx')
    ws_ideal_nadir =  wb["ideal_nadir"]
    ws_ref_point = wb["ref_point"]
    max_row = ws_ref_point.max_row
    if ref_point==None:
        st.header(performance_name[performance_measure])
        st.write(opt_type + " solution")
        st.write(best_fitness)
        if opt_type == 'best':
            ws_ideal_nadir['B'+str(performance_measure+2)] = best_fitness
        else:
            ws_ideal_nadir['C' + str(performance_measure + 2)] = best_fitness
    else:
        st.header('Best solution')
        st.write(best_fitness)
        ws_ref_point['A' + str(max_row + 1)] = str(reference_point_input)
        #ws_ref_point['A'+str(max_row+1)] = "["+str(reference_point_input[0]) +", " + str(reference_point_input[1]) +", "+ str(reference_point_input[2]) +"]"
        ws_ref_point['B' + str(max_row + 1)] = best_fitness
        ws_ref_point['C' + str(max_row + 1)] = str(best)

    print('best fitness slow \t\t', best_fitness)

    wb.save(project_name+'.xlsx')
    wb.close()
    #st.write(pop)
    #st.write(log)
    #st.stop()

def start_new_project(project_name, performance_name):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "ideal_nadir"
    for i in range(len(performance_name)):
        ws['A'+str(i+2)] = performance_name[i]
    ws['B1'] = 'ideal point'
    ws['C1'] = 'nadir point'
    ws1 = wb.create_sheet("ref_point")  # insert at the end (default)
    ws1['A1'] = 'reference point'
    ws1['B1'] = 'best fitness'
    ws1['C1'] = 'best function'
    wb.save(project_name+'.xlsx')
    wb.close()
    st.success('Your project was successfully created! Please follow the next steps.')

def verification(performance_measure, ref_point, reference_point_input, best_func, lambda_performance, nb_simulations):

    performance_name = ["Makespan", "Total tardiness", "Total waiting time"]
    def div(left, right):
        try:
            return left / right
        except ZeroDivisionError:
            return 1

    # define the functions and terminals
    pset = gp.PrimitiveSet("MAIN", 13)
    pset.addPrimitive(operator.add, 2)
    pset.addPrimitive(operator.sub, 2)
    pset.addPrimitive(operator.mul, 2)
    pset.addPrimitive(operator.neg, 1)
    pset.addPrimitive(min, 2)
    pset.addPrimitive(max, 2)
    pset.addPrimitive(div, 2)
    pset.addPrimitive(ifte, 3)
    # there is currently no constant value as function defined
    # maybe change here to random between 0 and 10 as mentioned in overleaf

    # Definitions of the abbreviations:
    # PT: Processing Time of an operation
    # RPT: Remaining processing time of a job
    # RNO: Remaining number of uncompleted operations of a job
    # DD: Due Date of a job
    # SPTQ: Sum of processing time in queue
    # APTQ: Average processing time in queue
    # MAXPTQ: Maximum processing time in queue
    # MINPTQ: Minimum processing time in queue
    # MAXDD: Maximum Due date of remaining jobs
    # NRJ: Number of remaining jobs
    # SPT: Sum of processing time of all waiting jobs in the problem
    # TRNO: Total remaining number of uncompleted operations in the problem
    # CT: Current time

    # rename the terminals
    pset.renameArguments(ARG0='PT')
    pset.renameArguments(ARG1='RPT')
    pset.renameArguments(ARG2='RNO')
    pset.renameArguments(ARG3='DD')
    pset.renameArguments(ARG4='SPTQ')
    pset.renameArguments(ARG5='APTQ')
    pset.renameArguments(ARG6='MAXPTQ')
    pset.renameArguments(ARG7='MINPTQ')
    pset.renameArguments(ARG8='MAXDDQ')
    pset.renameArguments(ARG9='NJQ')
    pset.renameArguments(ARG10='SPT')
    pset.renameArguments(ARG11='TRNO')
    pset.renameArguments(ARG12='CT')
    opt_type = 'best'
    if opt_type=="best":
        creator.create("FitnessMin", base.Fitness, weights=(-1.0,))
        creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMin)
    else:
        creator.create("FitnessMax", base.Fitness, weights=(1.0,))
        creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMax)

    toolbox = base.Toolbox()
    toolbox.register("expr", gp.genHalfAndHalf, pset=pset, min_=2, max_=6)
    toolbox.register("individual", tools.initIterate, creator.Individual, toolbox.expr)
    toolbox.register("population", tools.initRepeat, list, toolbox.individual)
    toolbox.register("compile", gp.compile, pset=pset)

    func = toolbox.compile(expr=best_func)

    random_seed = [0, 10, 15, 36, 27, 50, 82, 99, 103, 230] # train set
    #random_seed = [5, 20, 30, 42, 51, 65, 87, 101, 110, 242]  # test set
    #random_seed = [2, 8, 13, 38, 24, 53, 80, 94, 105, 229]  # test set 2
    results = []
    makespan = []
    tardiness = []
    waiting_time = []
    '''
    pool = mp.Pool(mp.cpu_count())
    results, makespan, tardiness, earliness = pool.starmap(simulation, [(15, 20, func, random_seed[simulations]) for simulations in range(nb_simulations)])
    pool.close()
    '''

    for simulations in range(nb_simulations):
        # perform simulation
        results_current_simulation_run, makespan_current_simulation_run, tardiness_current_simulation_run, waiting_time_current_simulation_run = \
            simulation(number_machines=15, number_jobs=20, func=func, random_seed=random_seed[simulations])
        # results.append(results_current_simulation_run)
        makespan.append(makespan_current_simulation_run)
        tardiness.append(tardiness_current_simulation_run)
        waiting_time.append(waiting_time_current_simulation_run)


    # calculate and print the performance measures after all simulation runs
    avg_makespan = np.mean(makespan)
    avg_total_tardiness = np.mean(tardiness)
    avg_waiting_time = np.mean(waiting_time)
    performance = [avg_makespan, avg_total_tardiness, avg_waiting_time]
    reference_point = reference_point_input
    roh = 0.0001
    # calculate the fitness value according to the achievement scalarizing function
    if ref_point == None:
        fitness = performance[performance_measure]
    else:
        fitness = max(lambda_performance[i] * (performance[i] - reference_point[i]) for i in
                      range(0, len(reference_point))) + roh * sum(
            lambda_performance[i] * (performance[i] - reference_point[i]) for i in range(0, len(reference_point)))

    results = results_current_simulation_run

    return fitness, results, performance

def verification_final(performance_measure, ref_point, reference_point_input, best_func, lambda_performance, nb_simulations):

    performance_name = ["Makespan", "Total tardiness", "Total waiting time"]
    def div(left, right):
        try:
            return left / right
        except ZeroDivisionError:
            return 1

    # define the functions and terminals
    pset = gp.PrimitiveSet("MAIN", 13)
    pset.addPrimitive(operator.add, 2)
    pset.addPrimitive(operator.sub, 2)
    pset.addPrimitive(operator.mul, 2)
    pset.addPrimitive(operator.neg, 1)
    pset.addPrimitive(min, 2)
    pset.addPrimitive(max, 2)
    pset.addPrimitive(div, 2)
    pset.addPrimitive(ifte, 3)
    # there is currently no constant value as function defined
    # maybe change here to random between 0 and 10 as mentioned in overleaf

    # Definitions of the abbreviations:
    # PT: Processing Time of an operation
    # RPT: Remaining processing time of a job
    # RNO: Remaining number of uncompleted operations of a job
    # DD: Due Date of a job
    # SPTQ: Sum of processing time in queue
    # APTQ: Average processing time in queue
    # MAXPTQ: Maximum processing time in queue
    # MINPTQ: Minimum processing time in queue
    # MAXDD: Maximum Due date of remaining jobs
    # NRJ: Number of remaining jobs
    # SPT: Sum of processing time of all waiting jobs in the problem
    # TRNO: Total remaining number of uncompleted operations in the problem
    # CT: Current time

    # rename the terminals
    pset.renameArguments(ARG0='PT')
    pset.renameArguments(ARG1='RPT')
    pset.renameArguments(ARG2='RNO')
    pset.renameArguments(ARG3='DD')
    pset.renameArguments(ARG4='SPTQ')
    pset.renameArguments(ARG5='APTQ')
    pset.renameArguments(ARG6='MAXPTQ')
    pset.renameArguments(ARG7='MINPTQ')
    pset.renameArguments(ARG8='MAXDDQ')
    pset.renameArguments(ARG9='NJQ')
    pset.renameArguments(ARG10='SPT')
    pset.renameArguments(ARG11='TRNO')
    pset.renameArguments(ARG12='CT')
    opt_type = 'best'
    if opt_type=="best":
        creator.create("FitnessMin", base.Fitness, weights=(-1.0,))
        creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMin)
    else:
        creator.create("FitnessMax", base.Fitness, weights=(1.0,))
        creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMax)

    toolbox = base.Toolbox()
    toolbox.register("expr", gp.genHalfAndHalf, pset=pset, min_=2, max_=6)
    toolbox.register("individual", tools.initIterate, creator.Individual, best_func)
    toolbox.register("population", tools.initRepeat, list, toolbox.individual)
    toolbox.register("compile", gp.compile, pset=pset)


    func = toolbox.compile(expr=best_func)

    random_seed = [0, 10, 15, 36, 27, 50, 82, 99, 103, 230] # train set
    #random_seed = [5, 20, 30, 42, 51, 65, 87, 101, 110, 242]  # test set
    #random_seed = [2, 8, 13, 38, 24, 53, 80, 94, 105, 229]  # test set 2
    results = []
    makespan = []
    tardiness = []
    waiting_time = []
    '''
    pool = mp.Pool(mp.cpu_count())
    results, makespan, tardiness, earliness = pool.starmap(simulation, [(15, 20, func, random_seed[simulations]) for simulations in range(nb_simulations)])
    pool.close()
    '''

    for simulations in range(nb_simulations):
        # perform simulation
        results_current_simulation_run, makespan_current_simulation_run, tardiness_current_simulation_run, waiting_time_current_simulation_run = \
            simulation(number_machines=15, number_jobs=20, func=func, random_seed=random_seed[simulations])
        # results.append(results_current_simulation_run)
        makespan.append(makespan_current_simulation_run)
        tardiness.append(tardiness_current_simulation_run)
        waiting_time.append(waiting_time_current_simulation_run)

    # calculate and print the performance measures after all simulation runs
    avg_makespan = np.mean(makespan)
    avg_total_tardiness = np.mean(tardiness)
    avg_waiting_time = np.mean(waiting_time)
    performance = [avg_makespan, avg_total_tardiness, avg_waiting_time]
    reference_point = reference_point_input
    roh = 0.0001
    # calculate the fitness value according to the achievement scalarizing function
    if ref_point == None:
        fitness = performance[performance_measure]
    else:
        fitness = max(lambda_performance[i] * (performance[i] - reference_point[i]) for i in
                      range(0, len(reference_point))) + roh * sum(
            lambda_performance[i] * (performance[i] - reference_point[i]) for i in range(0, len(reference_point)))

    results = results_current_simulation_run

    pop = toolbox.population(n=1)
    st.write(pop)

    return fitness, results, performance, func


def verification_verification(performance_measure, ref_point, reference_point_input, best_func):
    performance_name = ["makespan", "number tardy jobs", "total tardiness"]

    def div(left, right):
        try:
            return left / right
        except ZeroDivisionError:
            return 1

    pset = gp.PrimitiveSet("MAIN", 3)
    pset.addPrimitive(operator.add, 2)
    pset.addPrimitive(operator.sub, 2)
    pset.addPrimitive(operator.mul, 2)
    pset.addPrimitive(min, 2)
    pset.addPrimitive(max, 2)
    pset.addPrimitive(div, 2)
    # pset.addPrimitive(operator.neg, 1)
    # pset.addPrimitive(math.cos, 1)
    # pset.addPrimitive(math.sin, 1)
    # randomname = str(rd.randint(1, 100000000000000)) + str(rd.randint(1, 100000000000000)) + str(rd.randint(1, 100000000000000))
    # Problem bei Konstante da jedes mal neuer Name generiert werden muss
    # try:
    #    pset.addEphemeralConstant(randomname, lambda: rd.uniform(-1, 1))
    # except:
    #    randomname = str(rd.randint(1, 100000000000000)) + str(rd.randint(1, 100000000000000)) + str(
    #        rd.randint(1, 100000000000000))
    #    pset.addEphemeralConstant(randomname, lambda: rd.uniform(-1, 1))
    # maybe change here to random between 0 and 10 as mentioned in overleaf
    pset.renameArguments(ARG0='PT')
    pset.renameArguments(ARG1='SPT')
    pset.renameArguments(ARG2='RPT')
    opt_type = 'best'
    if opt_type == "best":
        creator.create("FitnessMin", base.Fitness, weights=(-1.0,))
        creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMin)
    else:
        creator.create("FitnessMax", base.Fitness, weights=(1.0,))
        creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMax)

    toolbox = base.Toolbox()
    toolbox.register("expr", gp.genHalfAndHalf, pset=pset, min_=2, max_=6)
    toolbox.register("individual", tools.initIterate, creator.Individual, toolbox.expr)
    toolbox.register("population", tools.initRepeat, list, toolbox.individual)
    toolbox.register("compile", gp.compile, pset=pset)

    func = toolbox.compile(expr=best_func)

    def visualize(results):
        schedule = pd.DataFrame(results)
        JOBS = sorted(list(schedule['Job'].unique()))
        MACHINES = sorted(list(schedule['Machine'].unique()))
        makespan = schedule['Finish'].max()

        bar_style = {'alpha': 1.0, 'lw': 25, 'solid_capstyle': 'butt'}
        text_style = {'color': 'white', 'weight': 'bold', 'ha': 'center', 'va': 'center'}
        colors = mpl.cm.Dark2.colors

        schedule.sort_values(by=['Job', 'Start'])
        schedule.set_index(['Job', 'Machine'], inplace=True)

        fig, ax = plt.subplots(2, 1, figsize=(12, 5 + (len(JOBS) + len(MACHINES)) / 4))

        for jdx, j in enumerate(JOBS, 1):
            for mdx, m in enumerate(MACHINES, 1):
                if (j, m) in schedule.index:
                    xs = schedule.loc[(j, m), 'Start']
                    xf = schedule.loc[(j, m), 'Finish']
                    ax[0].plot([xs, xf], [jdx] * 2, c=colors[mdx % 7], **bar_style)
                    ax[0].text((xs + xf) / 2, jdx, m, **text_style)
                    ax[1].plot([xs, xf], [mdx] * 2, c=colors[jdx % 7], **bar_style)
                    ax[1].text((xs + xf) / 2, mdx, j, **text_style)

        ax[0].set_title('Job Schedule')
        ax[0].set_ylabel('Job')
        ax[1].set_title('Machine Schedule')
        ax[1].set_ylabel('Machine')

        for idx, s in enumerate([JOBS, MACHINES]):
            ax[idx].set_ylim(0.5, len(s) + 0.5)
            ax[idx].set_yticks(range(1, 1 + len(s)))
            ax[idx].set_yticklabels(s)
            ax[idx].text(makespan, ax[idx].get_ylim()[0] - 0.2, "{0:0.1f}".format(makespan), ha='center', va='top')
            ax[idx].plot([makespan] * 2, ax[idx].get_ylim(), 'r--')
            ax[idx].set_xlabel('Time')
            ax[idx].grid(True)

        fig.tight_layout()
        # plt.show()
        st.pyplot(fig)

    def source(env, number, machine):
        """Source generates jobs randomly"""
        for i in range(number):
            processing_time = np.random.uniform(avg_processing_time - delta_processing_time,
                                                avg_processing_time + delta_processing_time)
            if i == 0:
                release_time = np.random.uniform(avg_job_interarrival_time - delta_release_date,
                                                 avg_job_interarrival_time + delta_release_date)
            else:
                release_time = release_time_list[i - 1] + np.random.uniform(
                    avg_job_interarrival_time - delta_release_date,
                    avg_job_interarrival_time + delta_release_date)

            due_date = release_time + k * processing_time
            due_date_list.append(due_date)
            processing_time_list.append(processing_time)
            release_time_list.append(release_time)
            c = job(env, f'Job {i + 1}', machine, processing_time=processing_time,
                    total_processing_time=total_processing_time)
            env.process(c)
            t = 0
            yield env.timeout(t)

    def job(env, name, machine, processing_time, total_processing_time, remaining_processing_time, due_date,
            release_date):
        """Job arrives, is served and leaves."""
        # arrive = release_time
        # print('%7.4f %s: Arrived' % (arrive, name))

        with machine.request(
                priority=func(processing_time, total_processing_time, remaining_processing_time, due_date)) as req:
            yield req
            # wait = env.now - arrive # waiting time of job
            job_start = env.now

            # We got to the counter
            # print('%7.4f %s: Waited %6.3f' % (env.now, name, wait))

            yield env.timeout(processing_time)
            # print('%7.4f %s: Finished' % (env.now, name))

            # Flow Time
            # flow_time.append(env.now-arrive)
            remaining_processing_time -= processing_time
            job_prec = name
            machine_prec = machine_list.index(machine)
            schedule.append([job_prec, machine_prec])
            results.append({'Job': f'J{job_prec}',
                            'Machine': f'M{machine_prec}',
                            'Start': job_start,
                            'Duration': processing_time,
                            'Finish': env.now})
            job_finished = 'Yes'
            for j in range(10):
                for m in range(10):
                    if TASKS[j, m]['prec'] == (job_prec, machine_prec):
                        machine = machine_list[m]
                        processing_time = TASKS[j, m]['dur']
                        job_finished = 'No'
                        env.process(
                            job(env, j, machine, processing_time, total_processing_time, remaining_processing_time,
                                due_date, release_date))
            if job_finished == 'Yes':
                # Completion time
                completion_time.append(env.now)
                # Tardiness of job
                tardiness.append(max(env.now - due_date, 0))
                # Tardy jobs
                if max(env.now - due_date, 0) > 0:
                    tardy_jobs.append(1)
                else:
                    tardy_jobs.append(0)

    number_simulations = 1
    avg_makespan_list = []
    avg_mean_tardiness_list = []
    avg_max_tardiness_list = []
    avg_total_tardiness_list = []
    avg_number_tardy_jobs_list = []

    for simulations in range(number_simulations):
        # TA15 benchmark problem
        number_machines, n_jobs = 15, 20
        data = """
         7 84  0 58 12 71  4 26  1 98  9 36  2 12 11 30 10 87 14 95  5 45  6 28 13 73  3 73  8 45 
         4 29  8 22  7 47  3 75  9 94 13 15 12  4  0 82 11 14 10 35  1 79  6 34  5 57 14 23  2 56 
         1 73  4 36  7 48 13 26  3 49  8 60 10 15  5 66 12 90 14 39  9  8  6 74  2 63  0 94 11 91 
         5  1 11 35  9 23 12 93  7 75  1 50  6 40 13 60  8 41  2  7  0 57 14 72  3 40  4 75 10  7 
         4 13 11 15 12 17  1 14  0 67  9 94  6 18 13 52  2 53 14 16  5 33 10 61  3 47  8 65  7 39 
         2 54  6 80  3 87  8 36 14 54  0 72  4 17 10 44 11 37  1 88  7 77 13 84 12 17  5 82  9 90 
         4  4 14 62  5 33 10 62  8 86  7 30  6 39  1 67  0 42 12 31  9 83 13 39 11 67  3 67  2 31 
         7 29 10 29 11 69 14 26  3 55  2 46  4 53  5 65  1 97 12 24  9 69  6 22 13 17  0 39  8 13 
        14 12 11 73  0 36 13 70  3 12  2 80  1 99  8 70  5 51  7 14  4 71 12 28  6 35 10 58  9 35 
         0 61  5 49 12 74  1 90 13 60 10 88  9  3  4 60  2 59  8 94 14 91 11 34  7 26  6  4  3 26 
         4 89  3 90  8 95 12 32  9 18 11 73  2  9 14 19  5 97  7 58 13 36  6 62 10 13  1 16  0  1 
         9 71  6 47  1 95  0  7 14 63  7 49 13 24 12 46  2 72 11 73  5 19  8 96 10 41  3 15  4 81 
         4 45  3  9  0 97 14 62 13 77  9 78  7 70  2 19 11 86  8 15 10 23  1 46  6 32 12  6  5 70 
        12 74 10 46  3 98  6  1  4 53  5 59  0 86  7 98  2 76  8 12 13 91 11 98 14 98  9 11  1 27 
        14 73  7 70  5 14  8 32 11 19  0 57  2 17 13 96 12 56  4 73  6 32  1  7 10 79  9 10  3 91 
         6 39 14 87 12 11  2 81  7  7  5 79  8 24 13  9 11 58  9 42  0 67  3 27  4 20  1 19 10 67 
         9 76  5 89 14 64 10 14 12 11  1 14  4 99 13 85  0 81 11  3  3 46  2 47  7 40  6 81  8 27 
         9 55 12 71  4  5 14 83 11 16  8  4  0 20  7 15  5 60  3  8  1 93 10 33  6 63 13 71  2 29 
        12 92  2 25  3  8 14 86  5 22  1 79  6 23 11 96 13 24  9 94  7 97 10 17  8 48  0 67  4 47 
         3  5 12 77 10 74  5 59 14 13  0 57  9 62  8 37 13 54  6 69 11 80  1 35  7 88  2 47  4 98 
             """
        TASKS = {}
        for job, line in enumerate(data.splitlines()[1:]):
            nums = line.split()
            prec = None
            for m, dur in zip(nums[::2], nums[1::2]):
                task = (job, int(m))
                TASKS[task] = {'dur': int(dur), 'prec': prec}
                prec = task

        release_date_list = []
        due_date_list_jobs = []
        total_processing_time = []
        # print(TASKS)

        for j in range(n_jobs):
            total_processing_time_current_job = 0
            for m in range(number_machines):
                total_processing_time_current_job += TASKS[j, m]['dur']
            total_processing_time.append(total_processing_time_current_job)
            due_date_list.append(1.3*total_processing_time_current_job)
        # print(TASKS['J0', 'M0']['prec'])
        # print(pd.DataFrame(TASKS).T)

        number_simulations = 1
        R_processing_time = 0.4
        avg_processing_time = 10
        processing_time = []
        duedate_tightness = 2
        duedate_variability = 0.3
        machine_utilization = 0.7
        job_interarrival_tightness = 1
        schedule = []
        release_time = []
        tardiness = []
        tardy_jobs = []
        completion_time = []
        flow_time = []
        due_date_list = []
        release_time_list = []
        processing_time_list = []
        results = []
        remaining_processing_time_current_job = []

        n_jobs_original = n_jobs
        delta_processing_time = R_processing_time * avg_processing_time
        avg_job_interarrival_time = avg_processing_time / machine_utilization
        delta_release_date = job_interarrival_tightness * avg_job_interarrival_time
        release_time.append(np.random.uniform(avg_job_interarrival_time - delta_release_date,
                                              avg_job_interarrival_time + delta_release_date))
        delta_duedate = duedate_tightness * duedate_variability
        k = np.random.uniform(duedate_tightness - delta_duedate, duedate_tightness + delta_duedate)

        env = simpy.Environment()

        # Start processes and run
        machine_0 = simpy.PriorityResource(env, capacity=1)
        machine_1 = simpy.PriorityResource(env, capacity=1)
        machine_2 = simpy.PriorityResource(env, capacity=1)
        machine_3 = simpy.PriorityResource(env, capacity=1)
        machine_4 = simpy.PriorityResource(env, capacity=1)
        machine_5 = simpy.PriorityResource(env, capacity=1)
        machine_6 = simpy.PriorityResource(env, capacity=1)
        machine_7 = simpy.PriorityResource(env, capacity=1)
        machine_8 = simpy.PriorityResource(env, capacity=1)
        machine_9 = simpy.PriorityResource(env, capacity=1)
        machine_10 = simpy.PriorityResource(env, capacity=1)
        machine_11 = simpy.PriorityResource(env, capacity=1)
        machine_12 = simpy.PriorityResource(env, capacity=1)
        machine_13 = simpy.PriorityResource(env, capacity=1)
        machine_14 = simpy.PriorityResource(env, capacity=1)
        # machine = simpy.Resource(env, capacity=1)

        machine_list = [machine_0, machine_1, machine_2, machine_3, machine_4, machine_5, machine_6, machine_7,
                        machine_8, machine_9, machine_10, machine_11, machine_12, machine_13, machine_14]

        for j in range(n_jobs):
            for m in range(number_machines):
                if TASKS[j, m]['prec'] == None:
                    current_machine = machine_list[m]
                    processing_time_new = float(TASKS[j, m]['dur'])
                    job_new = j
                    total_processing_time_current_job = total_processing_time[j]
                    remaining_processing_time_current_job = total_processing_time_current_job
                    release_date_current_job = release_date_list[j]
                    due_date_current_job = due_date_list_jobs[j]
                    env.process(job(env, job_new, current_machine, processing_time=processing_time_new,
                                    total_processing_time=total_processing_time_current_job,
                                    remaining_processing_time=remaining_processing_time_current_job,
                                    due_date=due_date_current_job, release_date=release_date_current_job))

        env.run()

        # for i in range(number_simulations):
        #    env.process(source(env, n_jobs, machine_1))
        #    env.run()

        # Post processing
        # calculate and performance measures of the current simulation run
        total_tardiness = sum(tardiness)
        mean_tardiness = mean(tardiness)
        max_tardiness = max(tardiness)
        number_tardy_jobs = sum(tardy_jobs)
        makespan = max(completion_time)
        # mean_flow_time = mean(flow_time)
        # max_flow_time = max(flow_time)
        # print('Release Time')
        # print(release_time_list)
        # print('processing time')
        # print(processing_time_list)
        # print('Due Dates')
        # print(due_date_list)
        # print(f'Total Tardiness: {total_tardiness}')
        # print(f'Mean Tardiness: {mean_tardiness}')
        # print(f'Max Tardiness: {max_tardiness}')
        # print(f'Number of tardy Jobs: {number_tardy_jobs}')
        # print(completion_time)
        # print(f'Makespan: {makespan}')
        # print(f'Mean flow time: {mean_flow_time}')
        # print(f'Max flow time: {max_flow_time}')
        # print(results)

        # add performance measures of current simulation run to the list for all runs
        avg_makespan_list.append(makespan)
        avg_mean_tardiness_list.append(mean_tardiness)
        avg_max_tardiness_list.append(max_tardiness)
        avg_total_tardiness_list.append(total_tardiness)
        avg_number_tardy_jobs_list.append(number_tardy_jobs)

        # visualize(results)

    # calculate and print the performance measures after all simulation runs
    avg_makespan = mean(avg_makespan_list)
    avg_mean_tardiness = mean(avg_mean_tardiness_list)
    avg_max_tardiness = mean(avg_max_tardiness_list)
    avg_total_tardiness = mean(avg_total_tardiness_list)
    avg_number_tardy_jobs = mean(avg_number_tardy_jobs_list)
    performance = [avg_makespan, avg_number_tardy_jobs, avg_total_tardiness]
    lambda_performance = [0.33, 0.33, 0.33]
    reference_point = reference_point_input
    roh = 0.0001
    fitness = max(lambda_performance[i] * (performance[i] - reference_point[i]) for i in
                  range(0, len(reference_point))) + roh * sum(
        lambda_performance[i] * (performance[i] - reference_point[i]) for i in range(0, len(reference_point)))
    # calculate the fitness value according to the achievement scalarizing function
    if ref_point == None:
        fitness = performance[performance_measure]
    else:
        fitness = max(lambda_performance[i] * (performance[i] - reference_point[i]) for i in
                      range(0, len(reference_point))) + roh * sum(
            lambda_performance[i] * (performance[i] - reference_point[i]) for i in range(0, len(reference_point)))

    # postprocessing
    st.header('Performance measures of the best solution')
    for i in range(len(performance_name)):
        st.subheader(performance_name[i])
        st.subheader(performance[i])
    st.header('Fitness')
    st.write(fitness)
    visualize(results)

def compute_reference_points(objective_ideal, objective_nadir, parameter_ref_point):
    # Calculate the reference point according to the sampling procedure
    reference_points = pd.DataFrame(
        [[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
         [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]])
    reference_points[0] = [objective_nadir[0], objective_ideal[1], objective_ideal[2]]
    reference_points[4] = [objective_ideal[0], objective_nadir[1], objective_ideal[2]]
    reference_points[14] = [objective_ideal[0], objective_ideal[1], objective_nadir[2]]
    # Calculate the reference point according to the sampling procedure / 40%
    reference_points_new = pd.DataFrame(
        [[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
         [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]])
    #parameter_ref_point = pd.Series([parameter_ref_point, parameter_ref_point, parameter_ref_point])
    reference_points_new[0] = [(parameter_ref_point * (objective_nadir[0]-objective_ideal[0]) + objective_ideal[0]), objective_ideal[1], objective_ideal[2]]
    reference_points_new[4] = [objective_ideal[0], parameter_ref_point*(objective_nadir[1]-objective_ideal[1]) + objective_ideal[1], objective_ideal[2]]
    reference_points_new[14] = [objective_ideal[0], objective_ideal[1], parameter_ref_point*(objective_nadir[2]-objective_ideal[2]) + objective_ideal[2]]
    # 0->4
    reference_points[1] = [
        reference_points.iloc[i, 0] + 1 / 4 * (reference_points.iloc[i, 4] - reference_points.iloc[i, 0]) for i in
        range(3)]
    reference_points[2] = [
        reference_points.iloc[i, 0] + 2 / 4 * (reference_points.iloc[i, 4] - reference_points.iloc[i, 0]) for i in
        range(3)]
    reference_points[3] = [
        reference_points.iloc[i, 0] + 3 / 4 * (reference_points.iloc[i, 4] - reference_points.iloc[i, 0]) for i in
        range(3)]
    # 0->4 / 40%
    reference_points_new[1] = [
        reference_points_new.iloc[i, 0] + 1 / 4 * (reference_points_new.iloc[i, 4] - reference_points_new.iloc[i, 0]) for i in
        range(3)]
    reference_points_new[2] = [
        reference_points_new.iloc[i, 0] + 2 / 4 * (reference_points_new.iloc[i, 4] - reference_points_new.iloc[i, 0]) for i in
        range(3)]
    reference_points_new[3] = [
        reference_points_new.iloc[i, 0] + 3 / 4 * (reference_points_new.iloc[i, 4] - reference_points_new.iloc[i, 0]) for i in
        range(3)]
    # 0->14
    reference_points[5] = [
        reference_points.iloc[i, 0] + 1 / 4 * (reference_points.iloc[i, 14] - reference_points.iloc[i, 0]) for i in
        range(3)]
    reference_points[9] = [
        reference_points.iloc[i, 0] + 2 / 4 * (reference_points.iloc[i, 14] - reference_points.iloc[i, 0]) for i in
        range(3)]
    reference_points[12] = [
        reference_points.iloc[i, 0] + 3 / 4 * (reference_points.iloc[i, 14] - reference_points.iloc[i, 0]) for i in
        range(3)]
    # 0->14 / 40%
    reference_points_new[5] = [
        reference_points_new.iloc[i, 0] + 1 / 4 * (reference_points_new.iloc[i, 14] - reference_points_new.iloc[i, 0]) for i in
        range(3)]
    reference_points_new[9] = [
        reference_points_new.iloc[i, 0] + 2 / 4 * (reference_points_new.iloc[i, 14] - reference_points_new.iloc[i, 0]) for i in
        range(3)]
    reference_points_new[12] = [
        reference_points_new.iloc[i, 0] + 3 / 4 * (reference_points_new.iloc[i, 14] - reference_points_new.iloc[i, 0]) for i in
        range(3)]
    # 4->14
    reference_points[8] = [
        reference_points.iloc[i, 4] + 1 / 4 * (reference_points.iloc[i, 14] - reference_points.iloc[i, 4]) for i in
        range(3)]
    reference_points[11] = [
        reference_points.iloc[i, 4] + 2 / 4 * (reference_points.iloc[i, 14] - reference_points.iloc[i, 4]) for i in
        range(3)]
    reference_points[13] = [
        reference_points.iloc[i, 4] + 3 / 4 * (reference_points.iloc[i, 14] - reference_points.iloc[i, 4]) for i in
        range(3)]
    # 4->14 / 40%
    reference_points_new[8] = [
        reference_points_new.iloc[i, 4] + 1 / 4 * (reference_points_new.iloc[i, 14] - reference_points_new.iloc[i, 4]) for i in
        range(3)]
    reference_points_new[11] = [
        reference_points_new.iloc[i, 4] + 2 / 4 * (reference_points_new.iloc[i, 14] - reference_points_new.iloc[i, 4]) for i in
        range(3)]
    reference_points_new[13] = [
        reference_points_new.iloc[i, 4] + 3 / 4 * (reference_points_new.iloc[i, 14] - reference_points_new.iloc[i, 4]) for i in
        range(3)]
    # 1->13
    reference_points[6] = [
        reference_points.iloc[i, 1] + 1 / 3 * (reference_points.iloc[i, 13] - reference_points.iloc[i, 1]) for i in
        range(3)]
    reference_points[10] = [
        reference_points.iloc[i, 1] + 2 / 3 * (reference_points.iloc[i, 13] - reference_points.iloc[i, 1]) for i in
        range(3)]
    # 1->13 / 40%
    reference_points_new[6] = [
        reference_points_new.iloc[i, 1] + 1 / 3 * (reference_points_new.iloc[i, 13] - reference_points_new.iloc[i, 1]) for i in
        range(3)]
    reference_points_new[10] = [
        reference_points_new.iloc[i, 1] + 2 / 3 * (reference_points_new.iloc[i, 13] - reference_points_new.iloc[i, 1]) for i in
        range(3)]
    # 2->11
    reference_points[7] = [
        reference_points.iloc[i, 2] + 1 / 2 * (reference_points.iloc[i, 11] - reference_points.iloc[i, 2]) for i in
        range(3)]
    # 2->11 / 40%
    reference_points_new[7] = [
        reference_points_new.iloc[i, 2] + 1 / 2 * (reference_points_new.iloc[i, 11] - reference_points_new.iloc[i, 2]) for i in
        range(3)]
    return reference_points, reference_points_new

def visualize_ref_points(reference_points, reference_points_new, objective_ideal, objective_nadir):
    # Visualize the ideal and nadir point in the objective space as well as the grid of uniformly distributed reference points
    reference_points_t = reference_points.T
    reference_points_t.rename(columns={0: 'Makespan', 1: 'Total tardiness', 2: 'Total waiting time'}, inplace=True)
    reference_points_new_t = reference_points_new.T
    reference_points_new_t.rename(columns={0: 'Makespan', 1: 'Total tardiness', 2: 'Total waiting time'}, inplace=True)
    reference_points_new_t.index += 1
    fig2 = plt.figure()
    ax1 = fig2.add_subplot(projection='3d')
    xs = reference_points_t['Makespan']
    ys = reference_points_t['Total tardiness']
    zs = reference_points_t['Total waiting time']
    ax1.scatter(xs, ys, zs, label='Reference Points')
    ax1.scatter(objective_nadir[0], objective_nadir[1], objective_nadir[2], label='Nadir Point (approx.)')
    ax1.scatter(objective_ideal[0], objective_ideal[1], objective_ideal[2], label='Ideal Point (approx.)')
    # lines connecting the ideal and nadir points of each objective value
    ax1.plot([objective_ideal[0], objective_ideal[0]], [objective_ideal[1], objective_ideal[1]],
             [objective_ideal[2], objective_nadir[2]], color='r', dashes=[6, 2], linewidth=1)
    ax1.plot([objective_ideal[0], objective_ideal[0]], [objective_ideal[1], objective_nadir[1]],
             [objective_ideal[2], objective_ideal[2]], color='r', dashes=[6, 2], linewidth=1)
    ax1.plot([objective_ideal[0], objective_nadir[0]], [objective_ideal[1], objective_ideal[1]],
             [objective_ideal[2], objective_ideal[2]], color='r', dashes=[6, 2], linewidth=1)
    ax1.plot([objective_nadir[0], objective_nadir[0]], [objective_ideal[1], objective_nadir[1]],
             [objective_ideal[2], objective_ideal[2]], color='r', dashes=[6, 2], linewidth=1)
    ax1.plot([objective_ideal[0], objective_nadir[0]], [objective_nadir[1], objective_nadir[1]],
             [objective_ideal[2], objective_ideal[2]], color='r', dashes=[6, 2], linewidth=1)
    ax1.plot([objective_nadir[0], objective_nadir[0]], [objective_nadir[1], objective_nadir[1]],
             [objective_ideal[2], objective_nadir[2]], color='r', dashes=[6, 2], linewidth=1)
    ax1.plot([objective_nadir[0], objective_nadir[0]], [objective_ideal[1], objective_ideal[1]],
             [objective_ideal[2], objective_nadir[2]], color='r', dashes=[6, 2], linewidth=1)
    ax1.plot([objective_ideal[0], objective_ideal[0]], [objective_nadir[1], objective_nadir[1]],
             [objective_ideal[2], objective_nadir[2]], color='r', dashes=[6, 2], linewidth=1)
    ax1.plot([objective_ideal[0], objective_ideal[0]], [objective_ideal[1], objective_nadir[1]],
             [objective_nadir[2], objective_nadir[2]], color='r', dashes=[6, 2], linewidth=1)
    ax1.plot([objective_nadir[0], objective_nadir[0]], [objective_ideal[1], objective_nadir[1]],
             [objective_nadir[2], objective_nadir[2]], color='r', dashes=[6, 2], linewidth=1)
    ax1.plot([objective_ideal[0], objective_nadir[0]], [objective_nadir[1], objective_nadir[1]],
             [objective_nadir[2], objective_nadir[2]], color='r', dashes=[6, 2], linewidth=1)
    ax1.plot([objective_ideal[0], objective_nadir[0]], [objective_ideal[1], objective_ideal[1]],
             [objective_nadir[2], objective_nadir[2]], color='r', dashes=[6, 2], linewidth=1)
    # lines connecting the reference points
    ax1.plot([reference_points_t.iloc[0, 0], reference_points_t.iloc[4, 0]],
             [reference_points_t.iloc[0, 1], reference_points_t.iloc[4, 1]],
             [reference_points_t.iloc[0, 2], reference_points_t.iloc[4, 2]], color='b', dashes=[2, 2], linewidth=1)
    ax1.plot([reference_points_t.iloc[0, 0], reference_points_t.iloc[14, 0]],
             [reference_points_t.iloc[0, 1], reference_points_t.iloc[14, 1]],
             [reference_points_t.iloc[0, 2], reference_points_t.iloc[14, 2]], color='b', dashes=[2, 2], linewidth=1)
    ax1.plot([reference_points_t.iloc[4, 0], reference_points_t.iloc[14, 0]],
             [reference_points_t.iloc[4, 1], reference_points_t.iloc[14, 1]],
             [reference_points_t.iloc[4, 2], reference_points_t.iloc[14, 2]], color='b', dashes=[2, 2], linewidth=1)
    ax1.plot([reference_points_t.iloc[1, 0], reference_points_t.iloc[13, 0]],
             [reference_points_t.iloc[1, 1], reference_points_t.iloc[13, 1]],
             [reference_points_t.iloc[1, 2], reference_points_t.iloc[13, 2]], color='b', dashes=[2, 2], linewidth=1)
    ax1.plot([reference_points_t.iloc[2, 0], reference_points_t.iloc[11, 0]],
             [reference_points_t.iloc[2, 1], reference_points_t.iloc[11, 1]],
             [reference_points_t.iloc[2, 2], reference_points_t.iloc[11, 2]], color='b', dashes=[2, 2], linewidth=1)
    ax1.plot([reference_points_t.iloc[3, 0], reference_points_t.iloc[8, 0]],
             [reference_points_t.iloc[3, 1], reference_points_t.iloc[8, 1]],
             [reference_points_t.iloc[3, 2], reference_points_t.iloc[8, 2]], color='b', dashes=[2, 2], linewidth=1)
    ax1.plot([reference_points_t.iloc[1, 0], reference_points_t.iloc[5, 0]],
             [reference_points_t.iloc[1, 1], reference_points_t.iloc[5, 1]],
             [reference_points_t.iloc[1, 2], reference_points_t.iloc[5, 2]], color='b', dashes=[2, 2], linewidth=1)
    ax1.plot([reference_points_t.iloc[2, 0], reference_points_t.iloc[9, 0]],
             [reference_points_t.iloc[2, 1], reference_points_t.iloc[9, 1]],
             [reference_points_t.iloc[2, 2], reference_points_t.iloc[9, 2]], color='b', dashes=[2, 2], linewidth=1)
    ax1.plot([reference_points_t.iloc[3, 0], reference_points_t.iloc[12, 0]],
             [reference_points_t.iloc[3, 1], reference_points_t.iloc[12, 1]],
             [reference_points_t.iloc[3, 2], reference_points_t.iloc[12, 2]], color='b', dashes=[2, 2], linewidth=1)
    # lines connecting the 40% reference points
    ax1.plot([reference_points_new_t.iloc[0, 0], reference_points_new_t.iloc[4, 0]],
             [reference_points_new_t.iloc[0, 1], reference_points_new_t.iloc[4, 1]],
             [reference_points_new_t.iloc[0, 2], reference_points_new_t.iloc[4, 2]], color='b', linewidth=1)
    ax1.plot([reference_points_new_t.iloc[0, 0], reference_points_new_t.iloc[14, 0]],
             [reference_points_new_t.iloc[0, 1], reference_points_new_t.iloc[14, 1]],
             [reference_points_new_t.iloc[0, 2], reference_points_new_t.iloc[14, 2]], color='b', linewidth=1)
    ax1.plot([reference_points_new_t.iloc[4, 0], reference_points_new_t.iloc[14, 0]],
             [reference_points_new_t.iloc[4, 1], reference_points_new_t.iloc[14, 1]],
             [reference_points_new_t.iloc[4, 2], reference_points_new_t.iloc[14, 2]], color='b', linewidth=1)
    ax1.plot([reference_points_new_t.iloc[1, 0], reference_points_new_t.iloc[13, 0]],
             [reference_points_new_t.iloc[1, 1], reference_points_new_t.iloc[13, 1]],
             [reference_points_new_t.iloc[1, 2], reference_points_new_t.iloc[13, 2]], color='b', linewidth=1)
    ax1.plot([reference_points_new_t.iloc[2, 0], reference_points_new_t.iloc[11, 0]],
             [reference_points_new_t.iloc[2, 1], reference_points_new_t.iloc[11, 1]],
             [reference_points_new_t.iloc[2, 2], reference_points_new_t.iloc[11, 2]], color='b', linewidth=1)
    ax1.plot([reference_points_new_t.iloc[3, 0], reference_points_new_t.iloc[8, 0]],
             [reference_points_new_t.iloc[3, 1], reference_points_new_t.iloc[8, 1]],
             [reference_points_new_t.iloc[3, 2], reference_points_new_t.iloc[8, 2]], color='b', linewidth=1)
    ax1.plot([reference_points_new_t.iloc[1, 0], reference_points_new_t.iloc[5, 0]],
             [reference_points_new_t.iloc[1, 1], reference_points_new_t.iloc[5, 1]],
             [reference_points_new_t.iloc[1, 2], reference_points_new_t.iloc[5, 2]], color='b', linewidth=1)
    ax1.plot([reference_points_new_t.iloc[2, 0], reference_points_new_t.iloc[9, 0]],
             [reference_points_new_t.iloc[2, 1], reference_points_new_t.iloc[9, 1]],
             [reference_points_new_t.iloc[2, 2], reference_points_new_t.iloc[9, 2]], color='b', linewidth=1)
    ax1.plot([reference_points_new_t.iloc[3, 0], reference_points_new_t.iloc[12, 0]],
             [reference_points_new_t.iloc[3, 1], reference_points_new_t.iloc[12, 1]],
             [reference_points_new_t.iloc[3, 2], reference_points_new_t.iloc[12, 2]], color='b', linewidth=1)
    ax1.set_xlabel('Makespan', fontsize=8)
    ax1.set_ylabel('Total tardiness', fontsize=8)
    ax1.set_zlabel('Total waiting time', fontsize=8)
    ax1.tick_params(axis='both', which='major', labelsize=6)
    ax1.tick_params(axis='both', which='minor', labelsize=8)
    for i in range(15):  # plot each point + it's index as text above
        ax1.text(reference_points_new_t.iloc[i, 0], reference_points_new_t.iloc[i, 1],
                 reference_points_new_t.iloc[i, 2], reference_points_t.index[i] + 1, size=8, zorder=10,
                 color='k', verticalalignment='top')
    ax1.view_init(20, 15)
    ax1.legend(loc='best', bbox_to_anchor=(1, 1), fontsize='x-small')
    st.pyplot(fig2)
    # show the reference points in tabular form
    st.subheader('Reference Points (Table)')
    st.dataframe(reference_points_new_t, height=2200)

def initialization_table(project_name):
    try:
        initialization_table = pd.read_excel(project_name + '.xlsx', header=0,
                                             index_col=0)
        st.success('Your project was successfully loaded! Please follow the next steps.')
    except:
        st.error('Oops: the project could not be found! Please check the entered name or create a new project first')
        st.stop()
    return initialization_table

def calculate_ref_point_parameter(choice_ref_point):
    if choice_ref_point == 'ideal point':
        parameter_ref_point = 0
    else:
        if choice_ref_point == 'nadir point':
            parameter_ref_point = 2
        else:
            parameter_ref_point = choice_ref_point / 0.5
    return parameter_ref_point

def objective_ideal(initialization_table):
    objective_ideal = [initialization_table.at['Makespan', 'ideal point'],
                       initialization_table.at['Total tardiness', 'ideal point'],
                       initialization_table.at['Total waiting time', 'ideal point']]
    return objective_ideal

def objective_nadir(initialization_table):
    objective_nadir = [initialization_table.at['Makespan', 'nadir point'],
                       initialization_table.at['Total tardiness', 'nadir point'],
                       initialization_table.at['Total waiting time', 'nadir point']]
    return objective_nadir

def calculate_weightings_ASF(objective_ideal, objective_nadir):
    lambda_ASF = [1/(objective_nadir[0]-objective_ideal[0]), 1/(objective_nadir[1]-objective_ideal[1]), 1/(objective_nadir[2]-objective_ideal[2])]
    return lambda_ASF